import numpy as np
import joblib
from openpyxl import load_workbook
from collections import defaultdict
from keras.models import load_model
from tensorflow.keras.preprocessing import image
import os

# === Load Models ===
rf_model = joblib.load("model14.pkl")  # Alzheimer's numerical model
keras_model = load_model("parkinsons_model.keras")  # Parkinson's image model

# === Load Excel === and replace with your own fike path where city img is.
wb = load_workbook("/Users/sihaamkhalid/Desktop/personal_proj/STEM-FELLOWSHIP-COMP/city_img/Level1Results.xlsx", data_only=True)
ws = wb["Results"]

# === Extract data rows (skip header) ===
rows = list(ws.iter_rows(min_row=2, values_only=True))

# === Track predictions by player ===
player_risks = defaultdict(lambda: {"numerical": [], "image": []})

# === Image prediction function ===
def predict_image_risk(img_path):
    try:
        img = image.load_img(img_path, target_size=(224, 224), color_mode="grayscale")
        img_array = image.img_to_array(img)
        img_array = np.expand_dims(img_array, axis=0) / 255.0
        return keras_model.predict(img_array)[0][0]
    except Exception as e:
        print(f"❌ Image error at {img_path}: {e}")
        return None

# === Process each row ===
for row in rows:
    name = row[0]
    level = row[2]
    numerical_values = row[3:]  # Skip: [0] name, [1] image placeholder, [2] level

    if not name or not level:
        print("⚠️ Skipping row due to missing name or level.")
        continue

    try:
        # ==== Alzheimer's prediction (numerical) ====
        cleaned_values = [val if val is not None else 0 for val in numerical_values]
        X_numerical = np.array(cleaned_values).reshape(1, -1)
        numerical_pred = rf_model.predict_proba(X_numerical)[0][1]
        player_risks[name]["numerical"].append(numerical_pred)
        print(f"✅ Alzheimer's Risk for {name} (Level {level}): {numerical_pred:.3f}")

        # ==== Parkinson's prediction (image) ====
        img_path = f"/Users/sihaamkhalid/Desktop/personal_proj/STEM-FELLOWSHIP-COMP/city_img/{name}_level{level}.png"
        if os.path.exists(img_path):
            img_pred = predict_image_risk(img_path)
            if img_pred is not None:
                player_risks[name]["image"].append(img_pred)
                print(f"✅ Parkinson's Risk for {name} (Level {level}): {img_pred:.3f}")
        else:
            print(f"⚠️ Image not found for {name} Level {level}: {img_path}")

    except Exception as e:
        print(f"❌ Error processing {name}: {e}")

# === Final Risk Categorization ===
def categorize_risk(score):
    if score <= 0.33:
        return "Low Risk"
    elif score <= 0.66:
        return "Medium Risk"
    return "High Risk"

print("\n=== Final Risk Assessment by Patient ===")
for name, preds in player_risks.items():
    if not preds["numerical"] and not preds["image"]:
        continue

    avg_numerical = np.mean(preds["numerical"]) if preds["numerical"] else None
    avg_image = np.mean(preds["image"]) if preds["image"] else None

    # Combine if both exist
    if avg_numerical is not None and avg_image is not None:
        final_avg = (avg_numerical + avg_image) / 2
    elif avg_numerical is not None:
        final_avg = avg_numerical
    elif avg_image is not None:
        final_avg = avg_image
    else:
        continue

    risk_level = categorize_risk(final_avg)

    print(f"\n🧠Patient: {name}")
    if avg_numerical is not None:
        print(f"Avg. Alzheimer’s Risk:  {avg_numerical:.3f}")
    if avg_image is not None:
        print(f"Avg. Parkinson’s Risk:  {avg_image:.3f}")
    print(f"Final Risk Score:       {final_avg:.3f} → {risk_level}")
