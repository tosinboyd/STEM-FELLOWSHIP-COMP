import sys
import math
import os
import numpy as np
import pandas as pd
import time
import random
import tempfile
import warnings

from PIL import Image, ImageQt
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from parkinsons_detector import ParkinsonsDetector
from drawing_metrics_logger import DrawingMetricsLogger

from PyQt5.QtWidgets import *
from PyQt5.QtGui import*
from PyQt5.QtCore import*
from PyQt5 import *

warnings.filterwarnings('ignore')


class CursorSelectionScreen(QWidget):
    def __init__(self, player_name):
        super().__init__()
        self.player_name = player_name
        self.selected_cursor_pixmap = None
        self.setWindowTitle("Choose Your Cursor")
        self.setMinimumSize(2000, 1100)
        self.resize(2000, 1100)

        # Set background color
        self.setStyleSheet("background-color: #2c3e50;")

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Title
        title = QLabel(f"Choose Your Cursor, {self.player_name}!")
        title.setFont(QFont("Arial", 32, QFont.Bold))
        title.setStyleSheet("color: white; margin: 20px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Cursor selection grid
        cursor_layout = QGridLayout()
        cursor_widget = QWidget()
        cursor_widget.setLayout(cursor_layout)

        cursor_files = ["Images/C1.png", "Images/C2.png", "Images/C3.png", "Images/C4.png", "Images/C5.png","Images/C6.png"]
        color_list = ["FF6B6B", "4ECDC4", "FFE66D", "95E1D3", "3498DB", "E9C9C4"]

        for i, cursor_file in enumerate(cursor_files):
            btn = QPushButton()
            btn.setFixedSize(400, 400)

            # Load and set icon
            icon = QIcon(cursor_file)
            if icon.isNull():
                print(f"Warning: '{cursor_file}' not found or could not be loaded.")
            btn.setIcon(icon)
            btn.setIconSize(QSize(350, 350))  # Size inside the button

            # Style the button
            btn.setStyleSheet(f"""
                QPushButton {{
                    border: 3px solid #34495e;
                    border-radius: 200px;
                    background-color: #{color_list[i]};
                }}
                QPushButton:hover {{
                    border: 5px solid #e74c3c;
                }}
                QPushButton:pressed {{
                    background-color: #2c3e50;
                }}
            """)

            btn.clicked.connect(lambda checked, file=cursor_file: self.select_cursor(file))

            row = i // 3
            col = i % 3
            cursor_layout.addWidget(btn, row, col)

        layout.addWidget(cursor_widget)

        # Next button
        self.next_btn = QPushButton("Next")
        self.next_btn.setFont(QFont("Arial", 18))
        self.next_btn.setFixedSize(200, 60)
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 40px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.next_btn.setEnabled(False)
        self.next_btn.clicked.connect(self.open_main_dialog)
        layout.addWidget(self.next_btn, alignment=Qt.AlignCenter)

        self.setLayout(layout)

    def select_cursor(self, cursor_file):
        # Load cursor pixmap
        pixmap = QPixmap(cursor_file)
        if pixmap.isNull():
            # Create fallback image
            pixmap = QPixmap(40, 40)
            pixmap.fill(Qt.black)
            print(f"Warning: '{cursor_file}' not found, using fallback cursor.")

        self.selected_cursor_pixmap = pixmap.scaled(50, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.next_btn.setEnabled(True)



    def proceed_to_game(self):
        self.setCursor(QCursor(self.selected_cursor_pixmap))

    def open_main_dialog(self):
        self.setCursor(QCursor(self.selected_cursor_pixmap))
        self.hide()
        self.dialog = QDialog()
        self.ui = Ui_Dialog()
        # Call setupUi with the required parameters
        self.ui.setupUi(self.dialog, self.player_name, self.selected_cursor_pixmap)
        self.dialog.exec_()


class Ui_Dialog(object):
    def setupUi(self, Dialog, player_name):
        self.Dialog = Dialog
        self.player_name = player_name

        Dialog.setObjectName("Dialog")
        Dialog.resize(2000, 1100)
        Dialog.setWindowTitle(f"Game Menu - Welcome {player_name}!")


        self.label = QLabel("Pick your level to start",Dialog)
        self.label.setFont(QFont("Fantasy"))
        self.label.setStyleSheet("color: white;font-size: 80px;text-align: center;font-weight: bold;background-color: #274e13;")
        self.label.setMinimumSize(2000,100)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.move(0,0)





        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(QtCore.QRect(120, 150, 400, 400))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #3498db;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)
        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setGeometry(QtCore.QRect(570, 150, 400, 400))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #6aa84f;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_3 = QtWidgets.QPushButton(Dialog)
        self.pushButton_3.setGeometry(QtCore.QRect(1020, 150, 400, 400))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_4 = QtWidgets.QPushButton(Dialog)
        self.pushButton_4.setGeometry(QtCore.QRect(1470, 150, 400, 400))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #e73c82;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_5 = QtWidgets.QPushButton(Dialog)
        self.pushButton_5.setGeometry(QtCore.QRect(120, 650, 400, 400))
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_5.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #a13ce7;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_6 = QtWidgets.QPushButton(Dialog)
        self.pushButton_6.setGeometry(QtCore.QRect(570, 650, 400, 400))
        self.pushButton_6.setObjectName("pushButton_6")
        self.pushButton_6.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #e7a13c;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_7 = QtWidgets.QPushButton(Dialog)
        self.pushButton_7.setGeometry(QtCore.QRect(1020, 650, 400, 400))
        self.pushButton_7.setObjectName("pushButton_7")
        self.pushButton_7.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #3d6d61;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.pushButton_8 = QtWidgets.QPushButton(Dialog)
        self.pushButton_8.setGeometry(QtCore.QRect(1470, 650, 400, 400))
        self.pushButton_8.setObjectName("pushButton_8")
        self.pushButton_8.setStyleSheet("""
            QPushButton {
                border-radius: 200px;
                background-color: #22299d;
                color: white;
                font-weight: bold;
                font-size: 70px;
            }
        """)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

        self.pushButton.clicked.connect(self.open_Level1Screen)
        self.pushButton_2.clicked.connect(self.open_Level2Screen)
        self.pushButton_3.clicked.connect(self.open_Level3Screen)
        self.pushButton_4.clicked.connect(self.open_Level4Screen)
        self.pushButton_5.clicked.connect(self.open_Level5Screen)
        self.pushButton_6.clicked.connect(self.open_Level6Screen)
        self.pushButton_7.clicked.connect(self.open_Level7Screen)
        self.pushButton_8.clicked.connect(self.open_Level8Screen)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", f"Main Menu - {self.player_name}"))
        self.pushButton.setText(_translate("Dialog", "Level 1"))
        self.pushButton_2.setText(_translate("Dialog", "Level 2"))
        self.pushButton_3.setText(_translate("Dialog", "Level 3"))
        self.pushButton_4.setText(_translate("Dialog", "Level 4"))
        self.pushButton_5.setText(_translate("Dialog", "Level 5"))
        self.pushButton_6.setText(_translate("Dialog", "Level 6"))
        self.pushButton_7.setText(_translate("Dialog", "Level 7"))
        self.pushButton_8.setText(_translate("Dialog", "Level 8"))




    def open_Level1Screen(self):
        self.pen_screen = Level1_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level2Screen(self):
        self.pen_screen = Level2_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level3Screen(self):
        self.pen_screen = Level3_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level4Screen(self):
        self.Dialog.hide()
        self.cursor_selection_screen = CursorSelectionScreen(self.player_name)

        def on_next_clicked():
            cursor_pixmap = self.cursor_selection_screen.selected_cursor_pixmap
            self.cursor_selection_screen.hide()
            self.pen_screen = Level4_Screen(self.Dialog, self.player_name, cursor_pixmap)
            self.pen_screen.show()

        self.cursor_selection_screen.next_btn.clicked.disconnect()
        self.cursor_selection_screen.next_btn.clicked.connect(on_next_clicked)
        self.cursor_selection_screen.show()
    def open_Level5Screen(self):
        self.pen_screen = Level5_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level6Screen(self):
        self.pen_screen = Level6_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level7Screen(self):
        self.pen_screen = Level7_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()
    def open_Level8Screen(self):
        self.pen_screen = Level8_Screen(self.Dialog, self.player_name)
        self.Dialog.hide()
        self.pen_screen.show()


class Level1_Screen(QWidget):
    def __init__(self, main_dialog, player_name):
        super().__init__()
        self.player_name = player_name
        self.main_dialog = main_dialog
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.setWindowTitle("Level 1")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)

        self.reference_layer = QPixmap(self.size())
        self.reference_layer.fill(Qt.transparent)

        self.drawing = QPixmap(self.size())
        self.drawing.fill(Qt.transparent)

        self.last_point = None
        self.last_time = None

        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.last_pen_up_time = None

        self.pressure_readings = []
        self.pen_positions = []
        self.pen_timestamps = []
        self.pendown_count = 0

        all_colors = [
            QColor(214, 45, 32),
            QColor(255, 167, 0),
            QColor(107, 210, 219),
            QColor(92, 184, 92),
            QColor(206, 73, 147),
            QColor(106, 13, 131),
            QColor(238, 93, 108)
        ]
        self.colors = random.sample(all_colors, 3)
        self.dot_pairs = []
        self.dot_radius = 90
        self.guideline_length = 1500

        self._setup_ui()
        self._generate_dot_positions()
        self._draw_dots_and_guide_lines()

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)

            X_scaled = self.alzheimers_scaler.transform(X)

            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def _setup_ui(self):
        layout = QVBoxLayout()

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
            QPushButton {
                color: black;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
            QPushButton:hover {
                background-color: lightcyan;
            }
        """)

        self.label = QLabel(f"Level 1 - Welcome {self.player_name}\nTrace the lines connecting dots of the same color")
        self.label.setFont(QFont("Arial", 24))
        self.label.setStyleSheet(
            "color: black; background-color: rgba(255, 255, 255, 200); padding: 10px; border-radius: 10px;")
        self.label.setWordWrap(True)
        layout.addWidget(self.label, alignment=Qt.AlignCenter)
        layout.addStretch()

        self.setLayout(layout)

    def _generate_dot_positions(self):
        button_area = self.next_btn.geometry()

        min_x = 100
        max_x = 1900
        min_y = 50
        max_y = 1500

        if button_area.isValid():
            button_margin = 100
            button_left = button_area.left() - button_margin
            button_right = button_area.right() + button_margin
            button_top = button_area.top() - button_margin
            button_bottom = button_area.bottom() + button_margin
        else:
            button_left = button_right = button_top = button_bottom = 0

        self.dot_pairs = []

        available_height = max_y - min_y
        vertical_spacing = 260

        available_width = 1000
        if available_width >= self.guideline_length:
            start_x = min_x + (available_width - self.guideline_length) // 2
            end_x = start_x + self.guideline_length
        else:
            start_x = min_x
            end_x = max_x

        for i, color in enumerate(self.colors):
            y_position = min_y + (i + 1) * vertical_spacing

            if (button_left <= end_x and button_right >= start_x and
                    button_top <= y_position <= button_bottom):
                if y_position < (button_top + button_bottom) // 2:
                    y_position = button_top - 50
                else:
                    y_position = button_bottom + 50

            y_position = max(min_y, min(y_position, max_y))

            dot1 = QPoint(int(start_x), int(y_position))
            dot2 = QPoint(int(end_x), int(y_position))

            self.dot_pairs.append({
                'color': color,
                'dots': [dot1, dot2]
            })

    def _draw_custom_dashed_line(self, painter, start_point, end_point, dash_length=80, gap_length=65):
        dx = end_point.x() - start_point.x()
        dy = end_point.y() - start_point.y()
        total_distance = math.sqrt(dx * dx + dy * dy)

        if total_distance == 0:
            return

        unit_x = dx / total_distance
        unit_y = dy / total_distance

        current_distance = 70
        dash_cycle = dash_length + gap_length

        while current_distance < total_distance:
            dash_start_x = start_point.x() + current_distance * unit_x
            dash_start_y = start_point.y() + current_distance * unit_y

            dash_end_distance = min(current_distance + dash_length, total_distance)
            dash_end_x = start_point.x() + dash_end_distance * unit_x
            dash_end_y = start_point.y() + dash_end_distance * unit_y

            painter.drawLine(QPoint(int(dash_start_x), int(dash_start_y)),
                             QPoint(int(dash_end_x), int(dash_end_y)))

            current_distance += dash_cycle

    def _draw_wave_line(self, painter, start_point, end_point, wave_amplitude=50, wave_frequency=1.3):
        dx = end_point.x() - start_point.x()
        dy = end_point.y() - start_point.y()
        total_distance = math.sqrt(dx * dx + dy * dy)

        if total_distance == 0:
            return

        wave_points = []
        num_points = int(total_distance // 60)

        for i in range(num_points + 1):
            t = i / num_points if num_points > 0 else 0
            base_x = start_point.x() + t * dx
            base_y = start_point.y() + t * dy

            spike_height = random.uniform(-wave_amplitude, wave_amplitude * wave_frequency)

            if dx != 0 or dy != 0:
                perp_x = -dy / total_distance
                perp_y = dx / total_distance
            else:
                perp_x = perp_y = 0

            wave_x = base_x + spike_height * perp_x
            wave_y = base_y + spike_height * perp_y

            wave_points.append(QPoint(int(wave_x), int(wave_y)))

        for i in range(len(wave_points) - 1):
            painter.drawLine(wave_points[i], wave_points[i + 1])

    def _draw_dots_and_guide_lines(self):
        painter = QPainter(self.reference_layer)
        painter.setRenderHint(QPainter.Antialiasing)

        for i, pair in enumerate(self.dot_pairs):
            color = pair['color']
            dots = pair['dots']

            pen = QPen(color, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)

            if i == 0:
                pen.setStyle(Qt.SolidLine)
                painter.setPen(pen)
                painter.drawLine(dots[0], dots[1])
            elif i == 1:
                self._draw_custom_dashed_line(painter, dots[0], dots[1], dash_length=80, gap_length=65)
            elif i == 2:
                self._draw_wave_line(painter, dots[0], dots[1], wave_amplitude=60, wave_frequency=1.5)

            painter.setPen(QPen(Qt.black, 4))
            painter.setBrush(QBrush(color))

            for dot in dots:
                painter.drawEllipse(dot.x() - self.dot_radius, dot.y() - self.dot_radius,
                                    self.dot_radius * 2, self.dot_radius * 2)

        painter.end()

    def is_in_drawing_area(self, pos):
        btn_rect = self.next_btn.geometry()
        return not btn_rect.contains(pos)

    def handle_next(self):
        self.end_time = time.time()
        if self.last_pen_up_time is not None and self.end_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        self.run_alzheimers_analysis()
        self.save_image_and_log()
        self.show_popup_with_home()

    def save_image_and_log(self):
        try:
            final_drawing = QPixmap(self.drawing.size())
            final_drawing.fill(Qt.white)

            painter = QPainter(final_drawing)
            painter.drawPixmap(0, 0, self.drawing)
            painter.end()

            metrics = self.metrics_logger.save_complete_session(
                drawing_pixmap=final_drawing,
                player_name=self.player_name,
                level="Level 1",
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count
            )
        except Exception as e:
            pass

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        home_button = QPushButton("Home")
        home_button.setFixedSize(200, 60)
        home_button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        home_button.clicked.connect(self.open_main_screen)
        home_button.clicked.connect(popup.close)

        layout.addWidget(home_button, alignment=Qt.AlignCenter)

        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.accept()
        self.open_main_screen()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), Qt.white)
        painter.drawPixmap(0, 0, self.reference_layer)
        painter.drawPixmap(0, 0, self.drawing)

    def resizeEvent(self, event):
        new_drawing = QPixmap(self.size())
        new_drawing.fill(Qt.transparent)
        QPainter(new_drawing).drawPixmap(0, 0, self.drawing)
        self.drawing = new_drawing

        new_reference = QPixmap(self.size())
        new_reference.fill(Qt.transparent)
        self.reference_layer = new_reference

        self._generate_dot_positions()
        self._draw_dots_and_guide_lines()

    def mousePressEvent(self, event):
        timestamp = time.time()
        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp

        if event.button() == Qt.LeftButton:
            if self.is_in_drawing_area(event.pos()):
                if self.last_pen_up_time is not None:
                    self.air_time += (timestamp - self.last_pen_up_time)
                    self.last_pen_up_time = None

                self.last_point = event.pos()
                self.last_time = timestamp
            else:
                self.last_point = None

    def mouseMoveEvent(self, event):
        if event.buttons() & Qt.LeftButton and self.last_point:
            timestamp = time.time()

            painter = QPainter(self.drawing)
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))

            delta_time = timestamp - self.last_time
            self.paper_time += delta_time
            self.last_time = timestamp

    def mouseReleaseEvent(self, event):
        if self.last_point is not None:
            self.last_point = None
            self.last_pen_up_time = time.time()

    def tabletEvent(self, event: QTabletEvent):
        timestamp = time.time()
        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp

        if event.type() == QTabletEvent.TabletPress:
            if self.is_in_drawing_area(event.pos()):
                if self.last_pen_up_time is not None:
                    self.air_time += (timestamp - self.last_pen_up_time)
                    self.last_pen_up_time = None

                self.last_point = event.pos()
                self.last_time = timestamp
                self.pendown_count += 1
            else:
                self.last_point = None

        elif event.type() == QTabletEvent.TabletMove and self.last_point is not None:
            painter = QPainter(self.drawing)
            pressure = event.pressure()
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(pressure)

            delta_time = timestamp - self.last_time
            self.paper_time += delta_time
            self.last_time = timestamp

        elif event.type() == QTabletEvent.TabletRelease:
            if self.last_point is not None:
                self.last_point = None
                self.last_pen_up_time = timestamp

        event.accept()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level2_Screen(QWidget):
    def __init__(self,main_dialog, player_name):
        super().__init__()
        self.player_name = player_name
        self.main_dialog = main_dialog
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")
        self.V_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
        self.parkinsons_detector = ParkinsonsDetector(
            model_path=os.path.join(self.V_FOLDER, 'best_parkinsons_model.keras'),
            image_size=(128, 128)
        )

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.prediction_result = None
        self.saved_image_path = None

        self.setWindowTitle("Level 2")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)

        cursor_pixmap = QPixmap("Images/C6.png")
        cursor_for_game = cursor_pixmap.scaled(200, 200, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
        self.setCursor(QCursor(cursor_for_game))

        self.background_layer = QPixmap(self.size())
        self.background_layer.fill(Qt.transparent)

        self.drawing_layer = QPixmap(self.size())
        self.drawing_layer.fill(Qt.transparent)

        self.last_point = None
        self.last_time = None

        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.last_pen_up_time = None

        self.pressure_readings = []
        self.pen_positions = []
        self.pen_timestamps = []
        self.pendown_count = 0

        self._setup_ui()
        self.draw_background()

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER=r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)

            X_scaled = self.alzheimers_scaler.transform(X)

            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def _setup_ui(self):
        layout = QVBoxLayout()

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
                    QPushButton {
                        color: white;
                        border: none;
                        border-radius: 30px;
                        font-weight: bold;
                        font-size: 50px;
                        background-color: rgba(70, 130, 180, 180);
                    }
                """)

        self.label = QLabel(f"Hi {self.player_name}, Drive your car on the hills")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: white;font-weight: bold;")
        layout.addWidget(self.label, alignment=Qt.AlignCenter)
        layout.addStretch()

        self.setLayout(layout)

    def is_in_drawing_area(self, pos):
        btn_rect = self.next_btn.geometry()
        return not btn_rect.contains(pos)

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def draw_background(self):
        painter = QPainter(self.background_layer)

        width = self.width()
        height = self.height()

        sky_gradient = QLinearGradient(0, 0, 0, 900)
        sky_gradient.setColorAt(0, QColor(135, 206, 250))
        sky_gradient.setColorAt(1, QColor(70, 130, 180))
        painter.fillRect(0, 0, width, 900, sky_gradient)

        painter.setBrush(QBrush(Qt.white))
        painter.setPen(Qt.NoPen)

        cloud_data = [
            (50, 410, 40), (100, 400, 35), (150, 380, 40), (200, 390, 45),
            (200, 210, 40), (250, 210, 35), (300, 210, 45), (350, 210, 45), (290, 170, 55),
            (580, 380, 40), (650, 360, 50), (700, 390, 35),
            (1050, 220, 40), (900, 220, 45), (950, 220, 38), (1000, 220, 42), (980, 170, 55),
            (1220, 360, 48), (1290, 340, 40), (1330, 360, 35),
            (1590, 180, 60), (1650, 170, 45), (1690, 190, 40),
            (1800, 400, 42), (1850, 395, 38), (1900, 375, 55),
        ]

        for x, y, radius in cloud_data:
            painter.drawEllipse(x - radius, y - radius, radius * 2, radius * 2)

        amplitude = height // 7
        mid_y = 700

        sine_points = []

        soil_color = QColor(139, 69, 19)
        painter.setBrush(QBrush(soil_color))
        painter.setPen(Qt.NoPen)

        soil_polygon = QPolygon()

        for x in range(0, width):
            y = mid_y - int(math.sin(x * 0.01) * amplitude)
            sine_points.append((x, y))
            soil_polygon.append(QPoint(x, y))

        soil_polygon.append(QPoint(width - 1, height))
        soil_polygon.append(QPoint(0, height))

        painter.drawPolygon(soil_polygon)

        pen = QPen(QColor(0, 100, 0), 50, Qt.SolidLine, Qt.RoundCap)
        painter.setPen(pen)

        prev_point = (0, mid_y)
        for x in range(1, width):
            y = mid_y - int(math.sin(x * 0.01) * amplitude)
            curr_point = (x, y)
            painter.drawLine(prev_point[0], prev_point[1], curr_point[0], curr_point[1])
            prev_point = curr_point

        painter.end()

    def create_drawing_pixmap(self):
        original_size = self.drawing_layer.size()
        crop_left = 0
        crop_right = 0
        crop_top = 50
        crop_bottom = 50

        new_width = max(100, original_size.width() - crop_left - crop_right)
        new_height = max(100, original_size.height() - crop_top - crop_bottom)

        pixmap = QPixmap(new_width, new_height)
        pixmap.fill(Qt.white)

        painter = QPainter(pixmap)
        pen = QPen()
        pen.setWidth(3)
        pen.setColor(Qt.black)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)

        if len(self.pen_positions) > 1:
            for i in range(1, len(self.pen_positions)):
                x1, y1 = self.pen_positions[i - 1]
                x2, y2 = self.pen_positions[i]

                x1_adj = x1 - crop_left
                x2_adj = x2 - crop_left
                y1_adj = y1 - crop_top
                y2_adj = y2 - crop_top

                if (0 <= x1_adj < new_width and 0 <= y1_adj < new_height and
                        0 <= x2_adj < new_width and 0 <= y2_adj < new_height):
                    painter.drawLine(x1_adj, y1_adj, x2_adj, y2_adj)

        painter.end()
        return pixmap

    def run_parkinsons_detection(self, image_path):
        if not image_path or not os.path.exists(image_path):
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'Image file not found for analysis'
            }
            return False

        try:
            prediction_result = self.parkinsons_detector.predict(image_path)

            if prediction_result and prediction_result['prediction_successful']:
                self.prediction_result = prediction_result
                return True
            else:
                self.prediction_result = {
                    'risk_level': 'Detection Failed',
                    'interpretation': 'Unable to analyze the drawing'
                }
                return False

        except Exception as e:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'An error occurred during analysis'
            }
            return False

    def save_basic_metrics(self):
        try:
            if self.end_time is None:
                self.end_time = time.time()

            metrics = self.metrics_logger.calculate_metrics(
                self.pen_positions,
                self.pen_timestamps,
                self.pressure_readings,
                self.start_time,
                self.end_time,
                self.air_time,
                self.paper_time,
                self.pendown_count
            )

            minimal_pixmap = QPixmap(100, 100)
            minimal_pixmap.fill(Qt.white)

            session_id = self.metrics_logger.get_next_session_id(self.player_name)
            processed_path, excel_img_path = self.metrics_logger.process_and_save_image(
                minimal_pixmap, self.player_name, session_id, "Level 2"
            )

            self.metrics_logger.save_to_excel(
                self.player_name,
                "Level 2",
                metrics,
                processed_path,
                excel_img_path,
                self.prediction_result
            )

        except Exception as e:
            pass

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        alzheimers_success = self.run_alzheimers_analysis()

        if self.has_drawing_content():
            self.save_image_and_log_complete_metrics()
        else:
            self.save_basic_metrics()
            self.prediction_result = {
                'risk_level': 'No Drawing',
                'interpretation': 'No drawing content was detected'
            }

        self.show_popup_with_home()

    def save_image_and_log_complete_metrics(self):
        if self.metrics_logger is None:
            return None

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return None

            pixmap = self.create_drawing_pixmap()

            temp_image_filename = f"temp_{self.player_name}_Level2.png"
            temp_image_path = os.path.join(self.SAVE_FOLDER, temp_image_filename)

            if not os.path.exists(self.SAVE_FOLDER):
                os.makedirs(self.SAVE_FOLDER)

            pixmap.save(temp_image_path)

            detection_success = self.run_parkinsons_detection(temp_image_path)

            metrics_result = self.metrics_logger.save_complete_session(
                drawing_pixmap=pixmap,
                player_name=self.player_name,
                level="Level 2",
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count,
                prediction_result=self.prediction_result
            )

            self.cleanup_processed_image(temp_image_path)

            return temp_image_path

        except Exception as e:
            self.save_basic_metrics()
            return None

    def cleanup_processed_image(self, image_path):
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                pass

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)

            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 25px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        if self.prediction_result:
            park_title = QLabel("Parkinson's Risk Assessment")
            park_title.setWordWrap(True)
            park_title.setAlignment(Qt.AlignCenter)
            park_title.setStyleSheet("""
                QLabel {
                    color: #4169E1;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_title)

            park_risk_label = QLabel(f"Risk Level: {self.prediction_result.get('risk_level', 'Unknown')}")
            park_risk_label.setWordWrap(True)
            park_risk_label.setAlignment(Qt.AlignCenter)
            park_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_risk_label)
            park_interpretation = QLabel(self.prediction_result.get('interpretation', 'No interpretation available'))
            park_interpretation.setWordWrap(True)
            park_interpretation.setAlignment(Qt.AlignCenter)
            park_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 25px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(park_interpretation)

        home_btn = QPushButton("Home")
        home_btn.setFixedSize(200, 60)
        home_btn.clicked.connect(lambda: self.return_home_from_popup(popup))
        home_btn.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 25px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        layout.addWidget(home_btn, alignment=Qt.AlignCenter)
        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.close()
        self.open_main_screen()

    def save_image_and_log(self):
        try:
            final_drawing = QPixmap(self.drawing_layer.size())
            final_drawing.fill(Qt.white)

            painter = QPainter(final_drawing)
            painter.setCompositionMode(QPainter.CompositionMode_SourceOver)

            drawing_image = self.drawing_layer.toImage()

            for y in range(drawing_image.height()):
                for x in range(drawing_image.width()):
                    pixel = drawing_image.pixel(x, y)
                    alpha = (pixel >> 24) & 0xFF

                    if alpha > 0:
                        painter.setPen(QPen(Qt.black))
                        painter.drawPoint(x, y)

            painter.end()

            metrics = self.metrics_logger.save_complete_session(
                drawing_pixmap=final_drawing,
                player_name=self.player_name,
                level="Level 2",
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count
            )
        except Exception as e:
            pass

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(0, 0, self.background_layer)
        painter.drawPixmap(0, 0, self.drawing_layer)

    def resizeEvent(self, event):
        new_background = QPixmap(self.size())
        new_background.fill(Qt.transparent)
        QPainter(new_background).drawPixmap(0, 0, self.background_layer)
        self.background_layer = new_background

        new_drawing = QPixmap(self.size())
        new_drawing.fill(Qt.transparent)
        QPainter(new_drawing).drawPixmap(0, 0, self.drawing_layer)
        self.drawing_layer = new_drawing

        self.draw_background()

    def mousePressEvent(self, event):
        timestamp = time.time()
        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp

        if event.button() == Qt.LeftButton:
            if self.is_in_drawing_area(event.pos()):
                if self.last_pen_up_time is not None:
                    self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

                self.last_point = event.pos()
                self.last_time = timestamp
            else:
                self.last_point = None

    def mouseMoveEvent(self, event):
        if event.buttons() & Qt.LeftButton and self.last_point:
            timestamp = time.time()

            painter = QPainter(self.drawing_layer)
            pen = QPen(Qt.black, 25, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(1.0)

            delta_time = timestamp - self.last_time
            self.paper_time += delta_time
            self.last_time = timestamp

    def mouseReleaseEvent(self, event):
        self.last_point = None
        self.last_pen_up_time = time.time()

    def tabletEvent(self, event: QTabletEvent):
        timestamp = time.time()
        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp

        if event.type() == QTabletEvent.TabletPress:
            if self.is_in_drawing_area(event.pos()):
                if self.last_pen_up_time is not None:
                    self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

                self.last_point = event.pos()
                self.last_time = timestamp
                self.pendown_count += 1
            else:
                self.last_point = None

        elif event.type() == QTabletEvent.TabletMove and self.last_point is not None:
            painter = QPainter(self.drawing_layer)
            pressure = event.pressure()
            pen = QPen(Qt.black,50, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(pressure)

            delta_time = timestamp - self.last_time
            self.paper_time += delta_time
            self.last_time = timestamp

        elif event.type() == QTabletEvent.TabletRelease:
            self.last_point = None
            self.last_time = timestamp
            self.last_pen_up_time = timestamp

        event.accept()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level3_Screen(QWidget):
    def __init__(self, main_dialog, player_name):
        super().__init__()
        self.setWindowTitle("Level 3 Cake Decorator")
        self.setFixedSize(2000, 1100)

        self.main_dialog = main_dialog
        self.player_name = player_name
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")

        self.V_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
        self.parkinsons_detector = ParkinsonsDetector(
            model_path=os.path.join(self.V_FOLDER, 'best_parkinsons_model.keras'),
            image_size=(128, 128)
        )

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.load_alzheimers_model()

        self.prediction_result = None
        self.saved_image_path = None

        self.last_point = None
        self.last_time = None

        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.last_pen_up_time = None

        self.pressure_readings = []
        self.pen_positions = []
        self.pen_timestamps = []
        self.pendown_count = 0

        self.cake_x, self.cake_y = 150, 310
        self.cake_width, self.cake_height = 1700, 700
        self.rotation_angle = 0
        self.drawing_strokes = []
        self.current_stroke = []
        self.is_drawing = False

        self.setup_ui()

    def load_alzheimers_model(self):
        try:
            import joblib
            model_folder = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(model_folder, "model14.pkl")
            scaler_path = os.path.join(model_folder, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                self.alzheimers_model = joblib.load(model_path)
                self.alzheimers_scaler = joblib.load(scaler_path)
        except:
            self.alzheimers_model = None
            self.alzheimers_scaler = None

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 20, 0, 0)
        self.setLayout(layout)

        self.label = QLabel(f"Hey {self.player_name}, follow the line and decorate the cake")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: black; font-weight: bold;")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label, alignment=Qt.AlignTop | Qt.AlignHCenter)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.setGeometry(1750, 950, 300, 150)
        self.next_btn.setStyleSheet(
            "color: black; border: none; border-radius: 30px; font-weight: bold; font-size: 50px;")
        self.next_btn.clicked.connect(self.handle_next)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(self.rect(), QColor(230, 240, 255))

        painter.save()
        painter.translate(self.cake_x + self.cake_width // 2, self.cake_y + self.cake_height // 2)
        painter.rotate(self.rotation_angle)
        painter.translate(-self.cake_width // 2, -self.cake_height // 2)

        self.draw_cake_base(painter)
        self.draw_candles(painter)
        self.draw_sine_wave_guide(painter)
        self.draw_user_drawings(painter)

        painter.restore()

    def draw_sine_wave_guide(self, painter):
        painter.setPen(QPen(QColor(255, 255, 255, 200), 6, Qt.DashLine))
        wave_amplitude, wave_frequency, y_center = 110, 5.5, self.cake_height // 2
        step = 20
        points = []
        for x in range(20, self.cake_width - 10, step):
            progress = (x - 20) / (self.cake_width - 10)
            angle = progress * wave_frequency * -2 * math.pi
            y = y_center + wave_amplitude * math.sin(angle)
            points.append(QPoint(x, int(y)))
        for i in range(len(points) - 1):
            painter.drawLine(points[i], points[i + 1])

    def draw_user_drawings(self, painter):
        painter.setPen(QPen(QColor(13, 190, 241), 30, Qt.SolidLine, Qt.RoundCap))
        for stroke in self.drawing_strokes:
            for i in range(len(stroke) - 1):
                painter.drawLine(stroke[i], stroke[i + 1])
        for i in range(len(self.current_stroke) - 1):
            painter.drawLine(self.current_stroke[i], self.current_stroke[i + 1])
    def draw_cake_base(self, painter):
        cake_color = QColor(139, 69, 19)
        painter.setBrush(QBrush(cake_color))
        painter.setPen(QPen(QColor(100, 50, 0), 2))
        cake_rect = QRectF(0, 5, self.cake_width, self.cake_height)
        painter.drawRoundedRect(cake_rect, 20, 20)

        painter.setBrush(QBrush(QColor(100, 50, 0)))
        painter.drawRoundedRect(QRectF(5, 10, self.cake_width - 10, self.cake_height - 10), 15, 15)

        painter.setBrush(QBrush(QColor(160, 82, 45)))
        painter.drawRoundedRect(QRectF(10, 15, self.cake_width - 20, self.cake_height - 20), 10, 10)

        painter.setBrush(QBrush(QColor(235, 169, 228)))
        painter.drawRoundedRect(QRectF(0, 0, self.cake_width, self.cake_height - 500), 20, 20)

        painter.setBrush(QBrush(QColor(125, 125, 125)))
        painter.drawRoundedRect(QRectF(-100, 700, 1900, 40), 20, 20)

    def draw_candles(self, painter):
        candle_count, candle_width, candle_height = 6, 50, 100
        candle_spacing = self.cake_width // (candle_count + 1)

        for i in range(candle_count):
            candle_x = candle_spacing * (i + 1) - candle_width // 2
            candle_y = -candle_height
            painter.setBrush(QBrush(QColor(241, 219, 13)))
            painter.setPen(QPen(QColor(200, 200, 100), 1))
            painter.drawRect(QRectF(candle_x, candle_y, candle_width, candle_height))

            flame_x, flame_y = candle_x + candle_width // 2, candle_y - 17
            flame_color = QColor(241, 124, 13)
            painter.setBrush(QBrush(flame_color))
            painter.setPen(QPen(QColor(255, 100, 0), 1))
            flame_points = [
                QPoint(flame_x, flame_y - 20), QPoint(flame_x + 8, flame_y + 10),
                QPoint(flame_x + 6, flame_y + 17), QPoint(flame_x, flame_y + 20),
                QPoint(flame_x - 6, flame_y + 17), QPoint(flame_x - 8, flame_y + 10),
            ]
            painter.drawPolygon(QPolygon(flame_points))

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            cake_point = self.screen_to_cake_coords(event.pos())
            if cake_point and self.point_in_cake(cake_point):
                current_time = time.time()

                if self.start_time is None:
                    self.start_time = current_time

                if self.last_pen_up_time is not None:
                    self.air_time += current_time - self.last_pen_up_time
                    self.last_pen_up_time = None

                self.is_drawing = True
                self.pendown_count += 1
                self.current_stroke = [cake_point]
                self.last_point = cake_point
                self.last_time = current_time

                self.pen_positions.append((cake_point.x(), cake_point.y()))
                self.pen_timestamps.append(current_time)
                self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))

                event.accept()

    def mouseMoveEvent(self, event):
        if self.is_drawing and event.buttons() & Qt.LeftButton:
            cake_point = self.screen_to_cake_coords(event.pos())
            if cake_point and self.point_in_cake(cake_point):
                current_time = time.time()
                self.current_stroke.append(cake_point)

                if self.last_time is not None:
                    self.paper_time += current_time - self.last_time

                self.pen_positions.append((cake_point.x(), cake_point.y()))
                self.pen_timestamps.append(current_time)
                self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))

                self.last_point = cake_point
                self.last_time = current_time

                self.update()
                event.accept()

    def mouseReleaseEvent(self, event):
        if self.is_drawing:
            current_time = time.time()

            if self.last_time is not None:
                self.paper_time += current_time - self.last_time

            if self.current_stroke:
                self.drawing_strokes.append(self.current_stroke.copy())
            self.current_stroke.clear()

            self.is_drawing = False
            self.last_pen_up_time = current_time

            event.accept()

    def screen_to_cake_coords(self, screen_point):
        center_x, center_y = self.cake_x + self.cake_width // 2, self.cake_y + self.cake_height // 2
        rel_x, rel_y = screen_point.x() - center_x, screen_point.y() - center_y
        angle_rad = math.radians(-self.rotation_angle)
        rotated_x = rel_x * math.cos(angle_rad) - rel_y * math.sin(angle_rad)
        rotated_y = rel_x * math.sin(angle_rad) + rel_y * math.cos(angle_rad)
        return QPoint(int(rotated_x + self.cake_width // 2), int(rotated_y + self.cake_height // 2))

    def point_in_cake(self, point):
        margin = 10
        return (margin <= point.x() <= self.cake_width - margin and
                margin <= point.y() <= self.cake_height - margin)

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np
            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0

            velocities = []
            for i in range(1, len(positions)):
                dt = timestamps[i] - timestamps[i - 1]
                if dt > 0:
                    dx = positions[i][0] - positions[i - 1][0]
                    dy = positions[i][1] - positions[i - 1][1]
                    velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                    velocities.append(velocity)

            accelerations = []
            for i in range(1, len(velocities)):
                dt = timestamps[i + 1] - timestamps[i]
                if dt > 0:
                    dv = velocities[i] - velocities[i - 1]
                    accelerations.append(dv / dt)

            jerks = []
            for i in range(1, len(accelerations)):
                dt = timestamps[i + 2] - timestamps[i + 1]
                if dt > 0:
                    da = accelerations[i] - accelerations[i - 1]
                    jerks.append(da / dt)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0
            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0
            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0
            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [total_time, self.air_time, self.paper_time, mean_speed, mean_acc,
                        pressure_mean, pressure_var, self.pendown_count, max_x, max_y,
                        gmrtp, mean_jerk, disp_index]

            return features
        except:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model or not self.alzheimers_scaler:
            return {'risk_score': 0.0, 'risk_level': 'Model Not Available',
                    'interpretation': 'Model unavailable'}

        try:
            import numpy as np
            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)
            X_scaled = self.alzheimers_scaler.transform(X)
            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)
            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "Low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "Moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "Higher risk. Professional evaluation recommended."

            return {'risk_score': float(risk_score), 'risk_level': risk_level,
                    'interpretation': interpretation}
        except:
            return {'risk_score': 0.0, 'risk_level': 'Prediction Error',
                    'interpretation': 'Error during analysis'}

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()
        if features is None:
            self.alzheimers_result = {'risk_score': 0.0, 'risk_level': 'No Data',
                                      'interpretation': 'Insufficient drawing data'}
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        if self.has_drawing_content():
            self.run_alzheimers_analysis()
            self.save_image_and_log_complete_metrics()
        else:
            self.save_basic_metrics()
            self.prediction_result = {
                'risk_level': 'No Drawing',
                'interpretation': 'No drawing content was detected'
            }

        self.show_popup_with_home()

    def create_drawing_pixmap(self):
        pixmap = QPixmap(self.cake_width, self.cake_height)
        pixmap.fill(Qt.white)

        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        pen = QPen()
        pen.setWidth(3)
        pen.setColor(Qt.black)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)

        if len(self.pen_positions) > 1:
            for i in range(1, len(self.pen_positions)):
                x1, y1 = self.pen_positions[i - 1]
                x2, y2 = self.pen_positions[i]
                painter.drawLine(x1, y1, x2, y2)

        painter.end()
        return pixmap

    def run_parkinsons_detection(self, image_path):
        if not image_path or not os.path.exists(image_path):
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'Image file not found for analysis'
            }
            return False

        try:
            prediction_result = self.parkinsons_detector.predict(image_path)

            if prediction_result and prediction_result['prediction_successful']:
                self.prediction_result = prediction_result
                return True
            else:
                self.prediction_result = {
                    'risk_level': 'Detection Failed',
                    'interpretation': 'Unable to analyze the drawing'
                }
                return False

        except Exception as e:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'An error occurred during analysis'
            }
            return False

    def save_basic_metrics(self):
        try:
            if self.end_time is None:
                self.end_time = time.time()

            metrics = self.metrics_logger.calculate_metrics(
                self.pen_positions,
                self.pen_timestamps,
                self.pressure_readings,
                self.start_time,
                self.end_time,
                self.air_time,
                self.paper_time,
                self.pendown_count
            )

            minimal_pixmap = QPixmap(100, 100)
            minimal_pixmap.fill(Qt.white)

            session_id = self.metrics_logger.get_next_session_id(self.player_name)
            processed_path, excel_img_path = self.metrics_logger.process_and_save_image(
                minimal_pixmap, self.player_name, session_id, "Level 3"
            )

            self.metrics_logger.save_to_excel(
                self.player_name,
                "Level 3",
                metrics,
                processed_path,
                excel_img_path,
                self.prediction_result
            )

        except Exception as e:
            pass

    def save_image_and_log_complete_metrics(self):
        if self.metrics_logger is None:
            return None

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return None

            pixmap = self.create_drawing_pixmap()

            temp_image_filename = f"temp_{self.player_name}_Level3.png"
            temp_image_path = os.path.join(self.SAVE_FOLDER, temp_image_filename)

            if not os.path.exists(self.SAVE_FOLDER):
                os.makedirs(self.SAVE_FOLDER)

            pixmap.save(temp_image_path)

            detection_success = self.run_parkinsons_detection(temp_image_path)

            metrics_result = self.metrics_logger.save_complete_session(
                drawing_pixmap=pixmap,
                player_name=self.player_name,
                level="Level 3",
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count,
                prediction_result=self.prediction_result
            )

            self.cleanup_processed_image(temp_image_path)

            return temp_image_path

        except Exception as e:
            self.save_basic_metrics()
            return None

    def cleanup_processed_image(self, image_path):
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                pass

    def show_popup_with_home(self):
        """Show the results popup with analysis results"""
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        # Title
        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        # Alzheimer's results
        if self.alzheimers_result:
            print(f"Displaying Alzheimer's results: {self.alzheimers_result}")

            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)
        else:
            print("No Alzheimer's results to display")

        # Parkinson's results
        if self.prediction_result:
            print(f"Displaying Parkinson's results: {self.prediction_result}")

            park_title = QLabel("Parkinson's Risk Assessment")
            park_title.setWordWrap(True)
            park_title.setAlignment(Qt.AlignCenter)
            park_title.setStyleSheet("""
                QLabel {
                    color: #4169E1;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_title)

            park_risk_label = QLabel(f"Risk Level: {self.prediction_result.get('risk_level', 'Unknown')}")
            park_risk_label.setWordWrap(True)
            park_risk_label.setAlignment(Qt.AlignCenter)
            park_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_risk_label)


            park_interpretation = QLabel(self.prediction_result.get('interpretation', 'No interpretation available'))
            park_interpretation.setWordWrap(True)
            park_interpretation.setAlignment(Qt.AlignCenter)
            park_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(park_interpretation)
        else:
            print("No Parkinson's results to display")


        home_button = QPushButton("Home")
        home_button.setFixedSize(200, 60)
        home_button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        home_button.clicked.connect(self.open_main_screen)
        home_button.clicked.connect(popup.close)

        layout.addWidget(home_button, alignment=Qt.AlignCenter)

        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.close()
        self.open_main_screen()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level4_Screen(QWidget):
    def __init__(self,main_dialog, player_name,cursor_pixmap):
        super().__init__()
        self.setWindowTitle("Level 4 - Wolfs Chase")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)

        self.main_dialog = main_dialog
        self.player_name = player_name
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        if not os.path.exists(self.SAVE_FOLDER):
            os.makedirs(self.SAVE_FOLDER)

        try:
            from drawing_metrics_logger import DrawingMetricsLogger
            self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")
        except ImportError:
            self.metrics_logger = None

        try:
            self.V_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            self.parkinsons_detector = ParkinsonsDetector(
                model_path=os.path.join(self.V_FOLDER, 'best_parkinsons_model.keras'),
                image_size=(128, 128)
            )
        except ImportError:
            self.parkinsons_detector = None

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.prediction_result = None
        self.saved_image_path = None

        self.last_point = None
        self.last_time = None
        self.last_pen_up_time = None

        self.pen_positions = []
        self.pen_timestamps = []
        self.pressure_readings = []
        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.pendown_count = 0
        self.is_drawing = False
        self.air_start_time = None
        self.paper_start_time = None

        self.drawing_strokes = []
        self.current_stroke = []

        self.layout = QVBoxLayout()
        home_btn = QPushButton("Home")
        home_btn.setFont(QFont("Arial", 16))
        home_btn.setFixedWidth(200)
        home_btn.clicked.connect(self.open_main_screen)
        self.layout.addWidget(home_btn, alignment=Qt.AlignLeft)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
                  QPushButton {
                      color: white;
                      border: none;
                      background-color: #0000FF;
                      border-radius: 30px;
                      font-weight: bold;
                      font-size: 50px;
                  }
              """)
        self.next_btn_rect = self.next_btn.geometry()

        self.label = QLabel(f"Hi {self.player_name}, get ready to run from the wolf")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: white;")
        self.layout.addWidget(self.label, alignment=Qt.AlignCenter)
        self.layout.addStretch()
        self.setLayout(self.layout)

        cursor_for_game = cursor_pixmap.scaled(200, 200, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
        self.setCursor(QCursor(cursor_for_game))

        self.drawing = QPixmap(2000, 1100)
        self.drawing.fill(Qt.darkGreen)

        self.trail_img = QPixmap("Images/wolf.png")
        if self.trail_img.isNull():
            self.trail_img = QPixmap(100, 100)
            self.trail_img.fill(Qt.red)

        self.trail_img = self.trail_img.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.trail_pos = QPoint(-100, -100)
        self.cursor_pos = QPoint(-100, -100)

        self.initial_trail_distance = 250
        self.final_trail_distance = 70
        self.current_trail_distance = self.initial_trail_distance

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_trail)
        self.timer.start(16)

        QTimer.singleShot(0, self.draw_spiral)

        if self.start_time is None:
            self.start_time = time.time()
            self.air_start_time = self.start_time

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER=r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)
            X_scaled = self.alzheimers_scaler.transform(X)
            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def draw_spiral(self):
        painter = QPainter(self.drawing)
        pen = QPen(Qt.black, 10)
        painter.setPen(pen)

        center_x = self.width() // 2
        center_y = 700
        a = 0
        b = 25
        theta = 0
        max_theta = 6 * math.pi
        step = 0.1

        prev_x, prev_y = center_x, center_y
        while theta < max_theta:
            r = a + b * theta
            x = center_x + int(r * math.cos(theta))
            y = center_y + int(r * math.sin(theta))
            painter.drawLine(prev_x, prev_y, x, y)
            prev_x, prev_y = x, y
            theta += step
        painter.end()
        self.update()

    def is_over_next_button(self, pos):
        self.next_btn_rect = self.next_btn.geometry()
        return self.next_btn_rect.contains(pos)

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(0, 0, self.drawing)

        trail_draw_pos = self.trail_pos - QPoint(self.trail_img.width() // 2, self.trail_img.height() // 2)
        painter.drawPixmap(trail_draw_pos, self.trail_img)

    def resizeEvent(self, event):
        new_pixmap = QPixmap(self.size())
        new_pixmap.fill(Qt.darkGreen)
        painter = QPainter(new_pixmap)
        painter.drawPixmap(0, 0, self.drawing)
        self.drawing = new_pixmap

    def mousePressEvent(self, event):
        if self.is_over_next_button(event.pos()):
            return

        if event.button() == Qt.LeftButton:
            current_time = time.time()

            if self.start_time is None:
                self.start_time = current_time
                self.last_time = current_time
                self.air_start_time = current_time

            if self.last_pen_up_time is not None:
                self.air_time += current_time - self.last_pen_up_time
                self.last_pen_up_time = None

            self.is_drawing = True
            self.pendown_count += 1
            self.current_stroke = [event.pos()]
            self.last_point = event.pos()
            self.paper_start_time = current_time
            self.last_time = current_time

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))

    def mouseMoveEvent(self, event):
        if self.is_over_next_button(event.pos()):
            return

        self.cursor_pos = event.pos()

        if self.is_drawing and event.buttons() & Qt.LeftButton and self.last_point:
            current_time = time.time()

            painter = QPainter(self.drawing)
            pen = QPen(Qt.darkMagenta, 15, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())

            self.current_stroke.append(event.pos())

            if self.last_time is not None:
                delta_time = current_time - self.last_time
                self.paper_time += delta_time

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(1.0)

            self.last_point = event.pos()
            self.last_time = current_time
            self.update()

    def mouseReleaseEvent(self, event):
        if self.is_drawing:
            current_time = time.time()

            if self.last_time is not None:
                self.paper_time += current_time - self.last_time

            if self.current_stroke:
                self.drawing_strokes.append(self.current_stroke.copy())
            self.current_stroke.clear()

            self.is_drawing = False
            self.last_pen_up_time = current_time
            self.last_point = None

    def tabletEvent(self, event: QTabletEvent):
        if self.is_over_next_button(event.pos()):
            return

        if event.type() == QTabletEvent.TabletPress:
            current_time = time.time()

            if self.start_time is None:
                self.start_time = current_time

            if self.last_pen_up_time is not None:
                self.air_time += current_time - self.last_pen_up_time
                self.last_pen_up_time = None

            self.is_drawing = True
            self.pendown_count += 1
            self.current_stroke = [event.pos()]
            self.last_point = event.pos()
            self.last_time = current_time

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(max(0.1, min(1.0, event.pressure())))

        elif event.type() == QTabletEvent.TabletMove and self.is_drawing and self.last_point:
            current_time = time.time()

            painter = QPainter(self.drawing)
            pen = QPen(Qt.darkMagenta, max(5, event.pressure() * 50), Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())

            self.current_stroke.append(event.pos())

            if self.last_time is not None:
                self.paper_time += current_time - self.last_time

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(max(0.1, min(1.0, event.pressure())))

            self.last_point = event.pos()
            self.last_time = current_time
            self.update()

        elif event.type() == QTabletEvent.TabletRelease:
            if self.is_drawing:
                current_time = time.time()

                if self.last_time is not None:
                    self.paper_time += current_time - self.last_time

                if self.current_stroke:
                    self.drawing_strokes.append(self.current_stroke.copy())
                self.current_stroke.clear()

                self.is_drawing = False
                self.last_pen_up_time = current_time
                self.last_point = None

        event.accept()

    def update_trail(self):
        if self.start_time is not None:
            transition_duration = 10.0
            elapsed_time = time.time() - self.start_time
            progress = min(1.0, elapsed_time / transition_duration)

            self.current_trail_distance = self.initial_trail_distance - (
                    self.initial_trail_distance - self.final_trail_distance) * progress

        dx = self.cursor_pos.x() - self.trail_pos.x()
        dy = self.cursor_pos.y() - self.trail_pos.y()

        current_distance = (dx * dx + dy * dy) ** 0.5

        if current_distance > 0:
            target_distance = self.current_trail_distance

            unit_dx = dx / current_distance
            unit_dy = dy / current_distance

            target_x = self.cursor_pos.x() - unit_dx * target_distance
            target_y = self.cursor_pos.y() - unit_dy * target_distance

            trail_speed = 0.1
            self.trail_pos += QPoint(
                int((target_x - self.trail_pos.x()) * trail_speed),
                int((target_y - self.trail_pos.y()) * trail_speed)
            )

        self.update()

    def create_drawing_pixmap(self):
        original_size = self.drawing.size()
        crop_left = 500
        crop_right = 450
        crop_top = 210
        crop_bottom = 50

        new_width = max(100, original_size.width() - crop_left - crop_right)
        new_height = max(100, original_size.height() - crop_top - crop_bottom)

        pixmap = QPixmap(new_width, new_height)
        pixmap.fill(Qt.white)

        painter = QPainter(pixmap)
        pen = QPen()
        pen.setWidth(3)
        pen.setColor(Qt.black)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)

        if len(self.pen_positions) > 1:
            for i in range(1, len(self.pen_positions)):
                x1, y1 = self.pen_positions[i - 1]
                x2, y2 = self.pen_positions[i]

                x1_adj = x1 - crop_left
                x2_adj = x2 - crop_left
                y1_adj = y1 - crop_top
                y2_adj = y2 - crop_top

                if (0 <= x1_adj < new_width and 0 <= y1_adj < new_height and
                        0 <= x2_adj < new_width and 0 <= y2_adj < new_height):
                    painter.drawLine(x1_adj, y1_adj, x2_adj, y2_adj)

        painter.end()
        return pixmap

    def run_parkinsons_detection(self, image_path):
        if not image_path or not os.path.exists(image_path) or not self.parkinsons_detector:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'Image file not found for analysis or detector not available'
            }
            return False

        try:
            prediction_result = self.parkinsons_detector.predict(image_path)

            if prediction_result and prediction_result.get('prediction_successful'):
                self.prediction_result = prediction_result
                return True
            else:
                self.prediction_result = {
                    'risk_level': 'Detection Failed',
                    'interpretation': 'Unable to analyze the drawing'
                }
                return False
        except Exception as e:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': f'An error occurred during analysis: {str(e)}'
            }
            return False

    def save_basic_metrics(self):
        if self.metrics_logger is None:
            return

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            total_time = self.end_time - self.start_time if self.start_time else 0

            blank_pixmap = QPixmap(100, 100)
            blank_pixmap.fill(Qt.white)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                self.metrics_logger.save_complete_session(
                    drawing_pixmap=blank_pixmap,
                    player_name=self.player_name,
                    level="Level 4",
                    pen_positions=[],
                    pen_timestamps=[],
                    pressure_readings=[],
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=total_time,
                    paper_time=0,
                    pendown_count=0,
                    prediction_result={
                        'risk_level': 'No Drawing',
                        'interpretation': 'No drawing content was detected'
                    }
                )
        except Exception as e:
            self.create_basic_excel_entry()

    def create_basic_excel_entry(self):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.metrics_logger.excel_path)
            ws = wb['Results']

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0

            row_data = [
                self.player_name, None, "Level 4", total_time, total_time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                'No Drawing', 'No drawing content was detected'
            ]

            ws.append(row_data)
            wb.save(self.metrics_logger.excel_path)
            wb.close()
        except Exception as e:
            pass

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        alzheimers_success = self.run_alzheimers_analysis()

        if self.has_drawing_content():
            self.save_image_and_log_complete_metrics()
        else:
            self.save_basic_metrics()
            self.prediction_result = {
                'risk_level': 'No Drawing',
                'interpretation': 'No drawing content was detected'
            }

        self.show_popup_with_home()

    def save_image_and_log_complete_metrics(self):
        if self.metrics_logger is None:
            return None

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return None

            pixmap = self.create_drawing_pixmap()

            temp_image_filename = f"temp_{self.player_name}_Level4.png"
            temp_image_path = os.path.join(self.SAVE_FOLDER, temp_image_filename)

            if not os.path.exists(self.SAVE_FOLDER):
                os.makedirs(self.SAVE_FOLDER)

            pixmap.save(temp_image_path)

            detection_success = self.run_parkinsons_detection(temp_image_path)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                metrics_result = self.metrics_logger.save_complete_session(
                    drawing_pixmap=pixmap,
                    player_name=self.player_name,
                    level="Level 4",
                    pen_positions=self.pen_positions,
                    pen_timestamps=self.pen_timestamps,
                    pressure_readings=self.pressure_readings,
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=self.air_time,
                    paper_time=self.paper_time,
                    pendown_count=self.pendown_count,
                    prediction_result=self.prediction_result
                )

            self.cleanup_processed_image(temp_image_path)

            return temp_image_path

        except Exception as e:
            self.save_basic_metrics()
            return None

    def cleanup_processed_image(self, image_path):
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                pass

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                        QLabel {
                            color: #444444;
                            font-size: 20px;
                            margin-bottom: 20px;
                            padding: 10px;
                            border: 1px solid #ddd;
                            border-radius: 5px;
                            background-color: #f9f9f9;
                        }
                    """)
            layout.addWidget(alz_interpretation)
        else:
            print("No Alzheimer's results to display")

            # Parkinson's results
        if self.prediction_result:
            print(f"Displaying Parkinson's results: {self.prediction_result}")

            park_title = QLabel("Parkinson's Risk Assessment")
            park_title.setWordWrap(True)
            park_title.setAlignment(Qt.AlignCenter)
            park_title.setStyleSheet("""
                        QLabel {
                            color: #4169E1;
                            font-weight: bold;
                            font-size: 25px;
                            margin-bottom: 5px;
                        }
                    """)
            layout.addWidget(park_title)

            park_risk_label = QLabel(f"Risk Level: {self.prediction_result.get('risk_level', 'Unknown')}")
            park_risk_label.setWordWrap(True)
            park_risk_label.setAlignment(Qt.AlignCenter)
            park_risk_label.setStyleSheet("""
                        QLabel {
                            color: Black;
                            font-weight: bold;
                            font-size: 25px;
                            margin-bottom: 5px;
                        }
                    """)
            layout.addWidget(park_risk_label)

            park_interpretation = QLabel(self.prediction_result.get('interpretation', 'No interpretation available'))
            park_interpretation.setWordWrap(True)
            park_interpretation.setAlignment(Qt.AlignCenter)
            park_interpretation.setStyleSheet("""
                        QLabel {
                            color: #444444;
                            font-size: 20px;
                            margin-bottom: 20px;
                            padding: 10px;
                            border: 1px solid #ddd;
                            border-radius: 5px;
                            background-color: #f9f9f9;
                        }
                    """)
            layout.addWidget(park_interpretation)
        else:
            print("No Parkinson's results to display")

        home_button = QPushButton("Home")
        home_button.setFixedSize(200, 60)
        home_button.setStyleSheet("""
                    QPushButton {
                        color: white;
                        border: none;
                        background-color: #4169E1;
                        border-radius: 15px;
                        font-weight: bold;
                        font-size: 18px;
                    }
                    QPushButton:hover {
                        background-color: #6495ED;
                    }
                """)
        home_button.clicked.connect(self.open_main_screen)
        home_button.clicked.connect(popup.close)

        layout.addWidget(home_button, alignment=Qt.AlignCenter)

        popup.setLayout(layout)
        popup.exec_()

    def categorize_combined_risk(self, score):
        if score <= 0.33:
            return "Low Risk"
        elif score <= 0.66:
            return "Medium Risk"
        else:
            return "High Risk"

    def return_home_from_popup(self, popup):
        popup.accept()
        self.open_main_screen()

    def is_in_drawing_area(self, pos):
        return 0 <= pos.x() < self.width() and 0 <= pos.y() < self.height()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level5_Screen(QWidget):
    def __init__(self,main_dialog, player_name):
        super().__init__()

        self.setWindowTitle("Level 5 - Snail Shell")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        if not os.path.exists(self.SAVE_FOLDER):
            os.makedirs(self.SAVE_FOLDER)

        try:
            from drawing_metrics_logger import DrawingMetricsLogger
            self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")
        except ImportError:
            self.metrics_logger = None

        try:
            self.V_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            self.parkinsons_detector = ParkinsonsDetector(
                model_path=os.path.join(self.V_FOLDER, 'best_parkinsons_model.keras'),
                image_size=(128, 128)
            )
        except ImportError:
            self.parkinsons_detector = None

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.main_dialog = main_dialog
        self.player_name = player_name
        self.last_point = None
        self.last_pen_up_time = None
        self.last_time = None

        self.pen_positions = []
        self.pen_timestamps = []
        self.pressure_readings = []
        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.pendown_count = 0
        self.is_drawing = False
        self.air_start_time = None
        self.paper_start_time = None

        self.prediction_result = None
        self.saved_image_path = None

        self.original_bg_pixmap = QPixmap("Images/snail_back1.png")
        if self.original_bg_pixmap.isNull():
            self.original_bg_pixmap = QPixmap(800, 600)
            self.original_bg_pixmap.fill(Qt.white)

        self.setup_ui()

        self.start_time = time.time()
        self.air_start_time = self.start_time

        self.drawing = QPixmap()
        self.reference_layer = QPixmap()

    def setup_ui(self):
        self.layout = QVBoxLayout()

        home_btn = QPushButton("Home")
        home_btn.setFont(QFont("Arial", 16))
        home_btn.setFixedWidth(200)
        home_btn.clicked.connect(self.open_main_screen)
        self.layout.addWidget(home_btn, alignment=Qt.AlignLeft)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #0000FF;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
        """)

        self.label = QLabel(f"Hey {self.player_name}, draw the spirals in the snail")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: black;")
        self.layout.addWidget(self.label, alignment=Qt.AlignCenter)
        self.layout.addStretch()
        self.setLayout(self.layout)

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER=r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)

            X_scaled = self.alzheimers_scaler.transform(X)

            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def showEvent(self, event):
        super().showEvent(event)

        self.drawing = QPixmap(self.size())
        self.drawing.fill(Qt.transparent)

        self.reference_layer = QPixmap(self.size())
        self.reference_layer.fill(Qt.transparent)

        self.draw_spiral()

    def create_white_background_drawing(self):
        original_size = self.drawing.size()
        crop_left = 700
        crop_right = 400
        crop_top = 200
        crop_bottom = 160

        new_width = max(100, original_size.width() - crop_left - crop_right)
        new_height = max(100, original_size.height() - crop_top - crop_bottom)

        white_drawing = QPixmap(new_width, new_height)
        white_drawing.fill(Qt.white)

        painter = QPainter(white_drawing)
        pen = QPen()
        pen.setWidth(3)
        pen.setColor(Qt.black)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)

        if len(self.pen_positions) > 1:
            for i in range(1, len(self.pen_positions)):
                x1, y1 = self.pen_positions[i - 1]
                x2, y2 = self.pen_positions[i]

                x1_adj = x1 - crop_left
                x2_adj = x2 - crop_left
                y1_adj = y1 - crop_top
                y2_adj = y2 - crop_top

                if (0 <= x1_adj < new_width and 0 <= y1_adj < new_height and
                        0 <= x2_adj < new_width and 0 <= y2_adj < new_height):
                    painter.drawLine(x1_adj, y1_adj, x2_adj, y2_adj)

        painter.end()
        return white_drawing

    def run_parkinsons_detection(self, image_path):
        if not image_path or not os.path.exists(image_path) or not self.parkinsons_detector:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': 'Image file not found for analysis or detector not available'
            }
            return False

        try:
            prediction_result = self.parkinsons_detector.predict(image_path)

            if prediction_result and prediction_result.get('prediction_successful'):
                self.prediction_result = prediction_result
                return True
            else:
                self.prediction_result = {
                    'risk_level': 'Detection Failed',
                    'interpretation': 'Unable to analyze the drawing'
                }
                return False
        except Exception as e:
            self.prediction_result = {
                'risk_level': 'Error',
                'interpretation': f'An error occurred during analysis: {str(e)}'
            }
            return False

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        alzheimers_success = self.run_alzheimers_analysis()

        if self.has_drawing_content():
            image_path = self.save_image_and_log_complete_metrics()
            if not image_path:
                self.prediction_result = {
                    'risk_level': 'Error',
                    'interpretation': 'Failed to process drawing for analysis'
                }
        else:
            self.save_basic_metrics()
            self.prediction_result = {
                'risk_level': 'No Drawing',
                'interpretation': 'No drawing content was detected'
            }

        self.show_popup_with_home()

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def save_image_and_log_complete_metrics(self):
        if self.metrics_logger is None:
            return None

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return None

            white_drawing = self.create_white_background_drawing()

            temp_image_filename = f"temp_{self.player_name}_Level5.png"
            temp_image_path = os.path.join(self.SAVE_FOLDER, temp_image_filename)

            white_drawing.save(temp_image_path)

            detection_success = self.run_parkinsons_detection(temp_image_path)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                metrics_result = self.metrics_logger.save_complete_session(
                    drawing_pixmap=white_drawing,
                    player_name=self.player_name,
                    level="Level 5",
                    pen_positions=self.pen_positions,
                    pen_timestamps=self.pen_timestamps,
                    pressure_readings=self.pressure_readings,
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=self.air_time,
                    paper_time=self.paper_time,
                    pendown_count=self.pendown_count,
                    prediction_result=self.prediction_result
                )

            self.cleanup_processed_image(temp_image_path)
            return temp_image_path

        except Exception as e:
            self.save_basic_metrics()
            return None

    def cleanup_processed_image(self, image_path):
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                pass

    def save_basic_metrics(self):
        if self.metrics_logger is None:
            return

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            total_time = self.end_time - self.start_time if self.start_time else 0

            blank_pixmap = QPixmap(100, 100)
            blank_pixmap.fill(Qt.white)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                self.metrics_logger.save_complete_session(
                    drawing_pixmap=blank_pixmap,
                    player_name=self.player_name,
                    level="Level 5",
                    pen_positions=[],
                    pen_timestamps=[],
                    pressure_readings=[],
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=total_time,
                    paper_time=0,
                    pendown_count=0,
                    prediction_result={
                        'risk_level': 'No Drawing',
                        'interpretation': 'No drawing content was detected'
                    }
                )
        except Exception as e:
            self.create_basic_excel_entry()

    def create_basic_excel_entry(self):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.metrics_logger.excel_path)
            ws = wb['Results']

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0

            row_data = [
                self.player_name, None, "Level 5", total_time, total_time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                'No Drawing', 'No drawing content was detected'
            ]

            ws.append(row_data)
            wb.save(self.metrics_logger.excel_path)
            wb.close()
        except Exception as e:
            pass

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        if self.prediction_result:
            park_title = QLabel("Parkinson's Risk Assessment")
            park_title.setWordWrap(True)
            park_title.setAlignment(Qt.AlignCenter)
            park_title.setStyleSheet("""
                QLabel {
                    color: #4169E1;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_title)

            park_risk_label = QLabel(f"Risk Level: {self.prediction_result.get('risk_level', 'Unknown')}")
            park_risk_label.setWordWrap(True)
            park_risk_label.setAlignment(Qt.AlignCenter)
            park_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(park_risk_label)

            park_interpretation = QLabel(self.prediction_result.get('interpretation', 'No interpretation available'))
            park_interpretation.setWordWrap(True)
            park_interpretation.setAlignment(Qt.AlignCenter)
            park_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(park_interpretation)

        home_button = QPushButton("Home")
        home_button.setFixedSize(200, 60)
        home_button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        home_button.clicked.connect(self.open_main_screen)
        home_button.clicked.connect(popup.close)

        layout.addWidget(home_button, alignment=Qt.AlignCenter)

        popup.setLayout(layout)
        popup.exec_()

    def draw_spiral(self):
        if self.reference_layer.isNull():
            return

        import math

        painter = QPainter(self.reference_layer)
        pen = QPen(Qt.black, 10)
        painter.setPen(pen)

        center_x = 1125
        center_y = 600
        a = 0
        b = 23
        theta = 0
        max_theta = 6 * math.pi
        step = 0.1

        prev_x, prev_y = center_x, center_y
        while theta < max_theta:
            r = a + b * theta
            x = center_x + int(r * math.cos(theta))
            y = center_y + int(r * math.sin(theta))
            painter.drawLine(prev_x, prev_y, x, y)
            prev_x, prev_y = x, y
            theta += step

        painter.end()
        self.update()

    def return_home_from_popup(self, popup):
        popup.accept()
        self.open_main_screen()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(200, 150, self.original_bg_pixmap)
        painter.drawPixmap(0, 0, self.reference_layer)
        painter.drawPixmap(0, 0, self.drawing)

    def resizeEvent(self, event):
        super().resizeEvent(event)

        size = event.size()
        if size.width() > 0 and size.height() > 0:
            self.drawing = QPixmap(size)
            self.drawing.fill(Qt.transparent)

            self.reference_layer = QPixmap(size)
            self.reference_layer.fill(Qt.transparent)

            self.draw_spiral()

    def mousePressEvent(self, event):
        timestamp = time.time()

        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp
            self.air_start_time = timestamp

        if event.button() == Qt.LeftButton:
            if self.is_in_drawing_area(event.pos()):
                if self.last_pen_up_time is not None:
                    self.air_time += (timestamp - self.last_pen_up_time)
                    self.last_pen_up_time = None

                self.is_drawing = True
                self.pendown_count += 1
                self.last_point = event.pos()
                self.paper_start_time = timestamp
                self.last_time = timestamp

                self.pen_positions.append((event.pos().x(), event.pos().y()))
                self.pen_timestamps.append(timestamp)
                self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))
            else:
                self.last_point = None
                self.is_drawing = False

    def mouseMoveEvent(self, event):
        if event.buttons() & Qt.LeftButton and self.last_point and self.is_drawing:
            timestamp = time.time()

            painter = QPainter(self.drawing)
            pen = QPen(Qt.darkBlue, 40, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(1.0)

            if self.last_time is not None:
                delta_time = timestamp - self.last_time
                self.paper_time += delta_time
            self.last_time = timestamp

    def mouseReleaseEvent(self, event):
        if self.is_drawing:
            self.is_drawing = False
            self.last_point = None
            self.last_pen_up_time = time.time()

    def is_in_drawing_area(self, pos):
        return 0 <= pos.x() < self.width() and 0 <= pos.y() < self.height()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level6_Screen(QWidget):
    def __init__(self, main_dialog, player_name):
        super().__init__()

        self.setWindowTitle("Level 6 - Name Writing")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        if not os.path.exists(self.SAVE_FOLDER):
            os.makedirs(self.SAVE_FOLDER)

        try:
            from drawing_metrics_logger import DrawingMetricsLogger
            self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")
        except ImportError:
            self.metrics_logger = None

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.main_dialog = main_dialog
        self.player_name = player_name
        self.last_point = None
        self.last_pen_up_time = None
        self.last_time = None

        self.pen_positions = []
        self.pen_timestamps = []
        self.pressure_readings = []
        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.pendown_count = 0
        self.is_drawing = False
        self.air_start_time = None
        self.paper_start_time = None

        self.saved_image_path = None

        self.drawing = None

        self.layout = QVBoxLayout()
        home_btn = QPushButton("Home")
        home_btn.setFont(QFont("Arial", 16))
        home_btn.setFixedWidth(200)
        home_btn.clicked.connect(self.open_main_screen)
        self.layout.addWidget(home_btn, alignment=Qt.AlignLeft)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #0000FF;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
        """)

        self.next_btn_rect = self.next_btn.geometry()

        self.label = QLabel(f"Hey {self.player_name}, let's try writing your name")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: white;")
        self.layout.addWidget(self.label, alignment=Qt.AlignCenter)
        self.layout.addStretch()
        self.setLayout(self.layout)

        self.start_time = time.time()
        self.air_start_time = self.start_time

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER=r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)
            X_scaled = self.alzheimers_scaler.transform(X)
            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def showEvent(self, event):
        super().showEvent(event)

        self.drawing = QPixmap(self.size())
        self.drawing.fill(Qt.darkCyan)

        self.draw_reference_name()

    def draw_reference_name(self):
        if self.drawing is None:
            return

        painter = QPainter(self.drawing)
        font = QFont("Lucida Handwriting", 100, QFont.Bold)
        font.setStyleStrategy(QFont.PreferAntialias)
        painter.setFont(font)
        painter.setPen(QPen(Qt.black, 3))

        metrics = painter.fontMetrics()
        text_width = metrics.horizontalAdvance(self.player_name)
        text_height = metrics.height()

        x = (self.width() - text_width) // 2
        y = (self.height() + text_height) // 2

        painter.drawText(x, y, self.player_name)
        painter.end()

    def is_over_next_button(self, pos):
        self.next_btn_rect = self.next_btn.geometry()
        return self.next_btn_rect.contains(pos)

    def create_white_background_drawing(self):
        if self.drawing is None:
            return QPixmap(self.size())

        original_height = self.drawing.height()
        original_width = self.drawing.width()
        crop_top = 300
        crop_bottom = 250

        if crop_top + crop_bottom >= original_height:
            crop_top = 0
            crop_bottom = 0

        cropped_height = original_height - crop_top - crop_bottom

        white_drawing = QPixmap(original_width, cropped_height)
        white_drawing.fill(Qt.white)

        painter = QPainter(white_drawing)
        pen = QPen()
        pen.setWidth(3)
        pen.setColor(Qt.black)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        painter.setPen(pen)

        if len(self.pen_positions) > 1:
            for i in range(1, len(self.pen_positions)):
                x1, y1 = self.pen_positions[i - 1]
                x2, y2 = self.pen_positions[i]

                y1_adj = y1 - crop_top
                y2_adj = y2 - crop_top

                if (0 <= y1_adj < cropped_height and 0 <= y2_adj < cropped_height):
                    painter.drawLine(x1, y1_adj, x2, y2_adj)

        painter.end()
        return white_drawing

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        alzheimers_success = self.run_alzheimers_analysis()

        if self.has_drawing_content():
            image_path = self.save_image_and_log_complete_metrics()
        else:
            self.save_basic_metrics()

        self.show_popup_with_home()

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def save_image_and_log_complete_metrics(self):
        if self.metrics_logger is None:
            return None

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return None

            white_drawing = self.create_white_background_drawing()

            import time as time_module
            timestamp = int(time_module.time())
            temp_image_filename = f"temp_{self.player_name}_Level6.png"
            temp_image_path = os.path.join(self.SAVE_FOLDER, temp_image_filename)

            white_drawing.save(temp_image_path)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                metrics_result = self.metrics_logger.save_complete_session(
                    drawing_pixmap=white_drawing,
                    player_name=self.player_name,
                    level="Level 6",
                    pen_positions=self.pen_positions,
                    pen_timestamps=self.pen_timestamps,
                    pressure_readings=self.pressure_readings,
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=self.air_time,
                    paper_time=self.paper_time,
                    pendown_count=self.pendown_count,
                    prediction_result=None
                )

            self.cleanup_processed_image(temp_image_path)
            return temp_image_path

        except Exception as e:
            self.save_basic_metrics()
            return None

    def cleanup_processed_image(self, image_path):
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                pass

    def save_basic_metrics(self):
        if self.metrics_logger is None:
            return

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            total_time = self.end_time - self.start_time if self.start_time else 0

            blank_pixmap = QPixmap(100, 100)
            blank_pixmap.fill(Qt.white)

            if hasattr(self.metrics_logger, 'save_complete_session'):
                self.metrics_logger.save_complete_session(
                    drawing_pixmap=blank_pixmap,
                    player_name=self.player_name,
                    level="Level 6",
                    pen_positions=[],
                    pen_timestamps=[],
                    pressure_readings=[],
                    start_time=self.start_time,
                    end_time=self.end_time,
                    air_time=total_time,
                    paper_time=0,
                    pendown_count=0,
                    prediction_result=None
                )
        except Exception as e:
            self.create_basic_excel_entry()

    def create_basic_excel_entry(self):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.metrics_logger.excel_path)
            ws = wb['Results']

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0

            row_data = [
                self.player_name, None, "Level 6", total_time, total_time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                'No Drawing', 'No drawing content was detected'
            ]

            ws.append(row_data)
            wb.save(self.metrics_logger.excel_path)
            wb.close()
        except Exception as e:
            pass

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        home_button = QPushButton("Home")
        home_button.setFixedSize(200, 60)
        home_button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        home_button.clicked.connect(self.open_main_screen)
        home_button.clicked.connect(popup.close)

        layout.addWidget(home_button, alignment=Qt.AlignCenter)

        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.accept()
        self.open_main_screen()

    def resizeEvent(self, event):
        super().resizeEvent(event)

        size = event.size()
        if size.width() > 0 and size.height() > 0:
            old_drawing = self.drawing
            self.drawing = QPixmap(size)
            self.drawing.fill(Qt.darkCyan)

            if old_drawing and not old_drawing.isNull():
                painter = QPainter(self.drawing)
                painter.drawPixmap(0, 0, old_drawing)
                painter.end()

            self.draw_reference_name()

    def paintEvent(self, event):
        if self.drawing is not None:
            base_painter = QPainter(self)
            base_painter.drawPixmap(0, 0, self.drawing)

    def mousePressEvent(self, event):
        if self.is_over_next_button(event.pos()):
            return

        timestamp = time.time()

        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp
            self.air_start_time = timestamp

        if event.button() == Qt.LeftButton:
            if self.last_pen_up_time is not None:
                self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

            self.is_drawing = True
            self.pendown_count += 1
            self.last_point = event.pos()
            self.paper_start_time = timestamp
            self.last_time = timestamp

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(max(0.1, min(1.0, 0.5 + np.random.normal(0, 0.1))))

    def mouseMoveEvent(self, event):
        if self.is_over_next_button(event.pos()):
            return

        if event.buttons() & Qt.LeftButton and self.last_point and self.is_drawing and self.drawing is not None:
            timestamp = time.time()

            painter = QPainter(self.drawing)
            pen = QPen(Qt.darkBlue, 25
                       , Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(1.0)

            if self.last_time is not None:
                delta_time = timestamp - self.last_time
                self.paper_time += delta_time
            self.last_time = timestamp

    def mouseReleaseEvent(self, event):
        if self.is_drawing:
            self.is_drawing = False
            self.last_point = None
            self.last_pen_up_time = time.time()

    def tabletEvent(self, event: QTabletEvent):
        if self.is_over_next_button(event.pos()):
            return

        timestamp = time.time()

        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp
            self.air_start_time = timestamp

        if event.type() == QTabletEvent.TabletPress:
            if self.last_pen_up_time is not None:
                self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

            self.is_drawing = True
            self.pendown_count += 1
            self.last_point = event.pos()
            self.paper_start_time = timestamp
            self.last_time = timestamp

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(event.pressure())

        elif event.type() == QTabletEvent.TabletMove and self.last_point and self.is_drawing and self.drawing is not None:
            painter = QPainter(self.drawing)
            pen = QPen(Qt.darkBlue, 25, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(event.pressure())

            if self.last_time is not None:
                delta_time = timestamp - self.last_time
                self.paper_time += delta_time
            self.last_time = timestamp

        elif event.type() == QTabletEvent.TabletRelease:
            if self.is_drawing:
                self.is_drawing = False
                self.last_point = None
                self.last_pen_up_time = time.time()

        event.accept()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()

class Level7_Screen(QWidget):
    def __init__(self, main_dialog, player_name):
        super().__init__()
        self.main_dialog = main_dialog
        self.player_name = player_name
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.load_alzheimers_model()

        self.setWindowTitle("Level 7 - Puppy Name Drawing")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)

        self.pen_positions = []
        self.pen_timestamps = []
        self.pressure_readings = []
        self.stroke_boundaries = []
        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.pendown_count = 0
        self.is_drawing = False
        self.air_start_time = None
        self.paper_start_time = None
        self.session_saved = False

        self.puppy_names = [
            "Falcon", "Jasper", "Magnus", "Poppy", "Teddy", "Biscuit", "Noodle",
            "Doodle", "Bandit", "Button", "Shadow", "Fuzzy", "Bubbles",
            "Sprinkle", "Cookie", "Tater Tot", "Waffles", "Peach", "Muffin",
            "Blossom", "Churro", "Fluffy", "Sunshine", "Honey", "Bailey",
            "Cupcake", "Daisy", "Chewy", "Pudding", "Pickles", "Wiggles",
            "Harley", "Gumdrop"
        ]

        self.current_puppy_name = ""
        self.last_point = None

        self.bg1_label = QLabel(self)
        self.bg1_label.setScaledContents(True)
        pixmap = QPixmap("Images/puppy.png")
        if not pixmap.isNull():
            self.bg1_label.setPixmap(pixmap)
            self.bg1_label.setGeometry(500, 100, 1000, 700)

        layout = QVBoxLayout()
        top_layout = QHBoxLayout()

        home_btn = QPushButton("Home")
        home_btn.setFixedWidth(200)
        home_btn.setFont(QFont("Arial", 16))
        home_btn.clicked.connect(self.open_main_screen)
        top_layout.addWidget(home_btn, alignment=Qt.AlignLeft)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
        """)

        self.name_btn = QPushButton("Puppy Names")
        self.name_btn.setFixedWidth(300)
        self.name_btn.setFont(QFont("Arial", 16))
        self.name_btn.clicked.connect(self.generate_new_name)
        top_layout.addWidget(self.name_btn, alignment=Qt.AlignRight)

        layout.addLayout(top_layout)

        self.label = QLabel(f"Hey {self.player_name}, Select and write the name for your Puppy")
        self.label.setFont(QFont("Arial", 28))
        self.label.setStyleSheet("color: white;")
        layout.addWidget(self.label, alignment=Qt.AlignCenter)
        layout.addStretch()
        self.setLayout(layout)

        self.drawing = QPixmap(self.size())
        self.drawing.fill(QColor(13, 190, 241))

        self.start_time = time.time()
        self.air_start_time = self.start_time

        self.draw_white_box()

    def load_alzheimers_model(self):
        try:
            import joblib
            model_folder = r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(model_folder, "model14.pkl")
            scaler_path = os.path.join(model_folder, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                self.alzheimers_model = joblib.load(model_path)
                self.alzheimers_scaler = joblib.load(scaler_path)
        except:
            self.alzheimers_model = None
            self.alzheimers_scaler = None

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np
            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0

            velocities = []
            for i in range(1, len(positions)):
                dt = timestamps[i] - timestamps[i - 1]
                if dt > 0:
                    dx = positions[i][0] - positions[i - 1][0]
                    dy = positions[i][1] - positions[i - 1][1]
                    velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                    velocities.append(velocity)

            accelerations = []
            for i in range(1, len(velocities)):
                dt = timestamps[i + 1] - timestamps[i]
                if dt > 0:
                    dv = velocities[i] - velocities[i - 1]
                    accelerations.append(dv / dt)

            jerks = []
            for i in range(1, len(accelerations)):
                dt = timestamps[i + 2] - timestamps[i + 1]
                if dt > 0:
                    da = accelerations[i] - accelerations[i - 1]
                    jerks.append(da / dt)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0
            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0
            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0
            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [total_time, self.air_time, self.paper_time, mean_speed, mean_acc,
                        pressure_mean, pressure_var, self.pendown_count, max_x, max_y,
                        gmrtp, mean_jerk, disp_index]

            return features
        except:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model or not self.alzheimers_scaler:
            return {'risk_score': 0.0, 'risk_level': 'Model Not Available',
                    'interpretation': 'Model unavailable'}

        try:
            import numpy as np
            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)
            X_scaled = self.alzheimers_scaler.transform(X)
            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)
            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "Low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "Moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "Higher risk. Professional evaluation recommended."

            return {'risk_score': float(risk_score), 'risk_level': risk_level,
                    'interpretation': interpretation}
        except:
            return {'risk_score': 0.0, 'risk_level': 'Prediction Error',
                    'interpretation': 'Error during analysis'}

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()
        if features is None:
            self.alzheimers_result = {'risk_score': 0.0, 'risk_level': 'No Data',
                                      'interpretation': 'Insufficient drawing data'}
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def draw_white_box(self):
        painter = QPainter(self.drawing)
        painter.setRenderHint(QPainter.Antialiasing)

        self.drawing.fill(QColor(13, 190, 241))

        box_width = 1200
        box_height = 340
        box_x = (self.width() - box_width) // 2
        box_y = 705

        painter.setBrush(QBrush(Qt.white))
        painter.setPen(QPen(Qt.black, 3))
        painter.drawRect(box_x, box_y, box_width, box_height)

        if self.current_puppy_name:
            font = QFont("Lucida Handwriting", 80, QFont.Bold)
            font.setStyleStrategy(QFont.PreferAntialias)
            painter.setFont(font)
            painter.setPen(QPen(QColor(157, 13, 241), 2))
            metrics = painter.fontMetrics()
            text_width = metrics.horizontalAdvance(self.current_puppy_name)
            text_height = metrics.height()
            text_x = box_x + (box_width - text_width) // 2
            text_y = box_y + (box_height + text_height) // 2 - 40
            painter.drawText(text_x, text_y, self.current_puppy_name)

        painter.end()
        self.update()

    def handle_next(self):
        self.pendown_count = self.pendown_count - 1
        if not self.session_saved:
            self.save_image_and_log()
        self.run_alzheimers_analysis()
        self.show_popup_with_home()

    def save_image_and_log(self):
        if not self.metrics_logger or self.session_saved:
            return

        try:
            self.end_time = time.time()

            if self.air_start_time and not self.is_drawing:
                self.air_time += self.end_time - self.air_start_time

            if self.paper_start_time and self.is_drawing:
                self.paper_time += self.end_time - self.paper_start_time

            final_drawing = self.create_final_image()

            metrics = self.metrics_logger.save_complete_session(
                drawing_pixmap=final_drawing,
                player_name=self.player_name,
                level="Level 7",
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count
            )
            self.session_saved = True

        except Exception as e:
            pass

    def create_final_image(self):
        box_width = 1200
        box_height = 340
        box_x = (self.width() - box_width) // 2
        box_y = 705

        final_pixmap = QPixmap(box_width, box_height)
        final_pixmap.fill(Qt.white)

        painter = QPainter(final_pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        if len(self.pen_positions) > 1:
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)

            stroke_boundaries = self.stroke_boundaries + [len(self.pen_positions)]

            for stroke_idx in range(len(stroke_boundaries) - 1):
                start_idx = stroke_boundaries[stroke_idx]
                end_idx = stroke_boundaries[stroke_idx + 1]

                for i in range(start_idx + 1, end_idx):
                    x1, y1 = self.pen_positions[i - 1]
                    x2, y2 = self.pen_positions[i]

                    if (box_x <= x1 <= box_x + box_width and box_y <= y1 <= box_y + box_height and
                            box_x <= x2 <= box_x + box_width and box_y <= y2 <= box_y + box_height):
                        rel_x1, rel_y1 = x1 - box_x, y1 - box_y
                        rel_x2, rel_y2 = x2 - box_x, y2 - box_y
                        painter.drawLine(rel_x1, rel_y1, rel_x2, rel_y2)

        painter.end()
        return final_pixmap

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Good Job!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        home_btn = QPushButton("Home")
        home_btn.setFixedWidth(300)
        home_btn.clicked.connect(lambda: self.return_home_from_popup(popup))
        home_btn.setStyleSheet("""
            QPushButton {
                background-color: lightblue;
                border-style: outset;
                border-width: 2px;
                border-radius: 30px;              
                border-color: beige;
                padding: 6px;
                color: Black;
                font-weight: bold;
                font-size: 40px;          
            }
        """)
        layout.addWidget(home_btn, alignment=Qt.AlignCenter)
        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.close()
        self.open_main_screen()

    def generate_new_name(self):
        available = [name for name in self.puppy_names if name != self.current_puppy_name]
        self.current_puppy_name = random.choice(available)
        self.draw_white_box()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(0, 0, self.drawing)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.drawing = QPixmap(self.size())
        self.drawing.fill(QColor(13, 190, 241))
        self.draw_white_box()

    def _start_drawing(self, pos, pressure=None):
        current_time = time.time()

        if not self.is_drawing:
            self.pendown_count += 1
            self.is_drawing = True

            self.stroke_boundaries.append(len(self.pen_positions))

            if self.air_start_time:
                self.air_time += current_time - self.air_start_time
            self.paper_start_time = current_time

        self.pen_positions.append((pos.x(), pos.y()))
        self.pen_timestamps.append(current_time)
        self.pressure_readings.append(pressure if pressure is not None else 0)

    def _end_drawing(self):
        if self.is_drawing:
            current_time = time.time()
            self.is_drawing = False

            if self.paper_start_time:
                self.paper_time += current_time - self.paper_start_time
            self.air_start_time = current_time

    def mousePressEvent(self, event):
        if self.next_btn.geometry().contains(event.pos()):
            return

        self.last_point = event.pos()
        self._start_drawing(event.pos())

    def mouseMoveEvent(self, event):
        if self.next_btn.geometry().contains(event.pos()):
            return

        if event.buttons() & Qt.LeftButton and self.last_point and self.is_drawing:
            painter = QPainter(self.drawing)
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            self.last_point = event.pos()

            current_time = time.time()
            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(0)

            self.update()

    def mouseReleaseEvent(self, event):
        self.last_point = None
        self._end_drawing()

    def tabletEvent(self, event: QTabletEvent):
        if self.next_btn.geometry().contains(event.pos()):
            return

        if event.type() == QTabletEvent.TabletPress:
            self.last_point = event.pos()
            self._start_drawing(event.pos(), pressure=event.pressure())

        elif event.type() == QTabletEvent.TabletMove and self.last_point and self.is_drawing:
            painter = QPainter(self.drawing)
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            self.last_point = event.pos()

            current_time = time.time()
            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(current_time)
            self.pressure_readings.append(event.pressure())

            self.update()

        elif event.type() == QTabletEvent.TabletRelease:
            self.last_point = None
            self._end_drawing()

        event.accept()

    def open_main_screen(self):
        if not self.session_saved:
            self.save_image_and_log()
        self.hide()
        self.main_dialog.show()

class Level8_Screen(QWidget):
    def __init__(self,main_dialog, player_name):
        super().__init__()

        self.setWindowTitle("Level 8 - Sentence Writing")
        self.setMinimumSize(2000, 1100)
        self.setMouseTracking(True)
        self.SAVE_FOLDER = r"C:\Users\Hooria\PycharmProjects\Project4\.venv\city_img"

        self.metrics_logger = DrawingMetricsLogger(self.SAVE_FOLDER, "Level1Results.xlsx")

        self.alzheimers_model = None
        self.alzheimers_scaler = None
        self.alzheimers_result = None
        self.alzheimers_model_loaded = False

        self.load_alzheimers_model()

        self.main_dialog = main_dialog
        self.player_name = player_name
        self.last_point = None
        self.last_pen_up_time = None
        self.last_time = None

        self.sentences = [
            "I brush my teeth.",
            "She makes her bed.",
            "We eat our food.",
            "He drinks coffee.",
            "I check my phone.",
            "We read the news.",
            "She takes a bath.",
            "He ties his shoes.",
            "I pack my bag.",
            "We go to work.",
            "He gets the mail.",
            "I open the door.",
            "They greet us.",
            "She locks the door.",
            "He buys the milk.",
            "We clean the room.",
            "We watch TV.",
            "She reads books.",
            "He plays games.",
            "I help cook.",
            "We eat dinner."
        ]
        self.current_sentence_index = 0

        self.pen_positions = []
        self.pen_timestamps = []
        self.pressure_readings = []
        self.start_time = None
        self.end_time = None
        self.air_time = 0
        self.paper_time = 0
        self.pendown_count = 0
        self.is_drawing = False
        self.air_start_time = None
        self.paper_start_time = None

        self.drawing = None
        self.user_drawing = None

        self.layout = QVBoxLayout()

        self.instruction_label = QLabel(f"Hey {self.player_name}, write the sentence below on the line:")
        self.instruction_label.setFont(QFont("Arial", 24))
        self.instruction_label.setStyleSheet("color: white;font-weight: bold;")
        self.layout.addWidget(self.instruction_label, alignment=Qt.AlignCenter)

        self.sentence_box = QLabel()
        self.sentence_box.setFixedHeight(120)
        self.sentence_box.setFixedWidth(800)
        self.sentence_box.setAlignment(Qt.AlignCenter)
        self.sentence_box.setStyleSheet("""
            QLabel {
                background-color: white;
                border: 3px solid black;
                border-radius: 15px;
                font-size: 60px;
                font-weight: bold;
                color: black;
                padding: 20px;
            }
        """)
        self.update_sentence_display()
        self.layout.addWidget(self.sentence_box, alignment=Qt.AlignCenter)

        self.change_sentence_btn = QPushButton("Change Sentence", self)
        self.change_sentence_btn.clicked.connect(self.change_sentence)
        self.change_sentence_btn.move(1500, 130)
        self.change_sentence_btn.resize(250, 80)
        self.change_sentence_btn.setStyleSheet("""
            QPushButton {
                background-color: #D7C8F4;
                color: black;
                border: 2px solid #D7C8F4;
                border-radius: 15px;
                font-weight: bold;
                font-size: 25px;
            }
            QPushButton:hover {
                background-color: 	#CAB8FF;
                color: white;
            }
        """)

        self.next_btn = QPushButton("Next", self)
        self.next_btn.clicked.connect(self.handle_next)
        self.next_btn.move(1750, 950)
        self.next_btn.resize(300, 150)
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #CAB8FF;
                color: black;
                border: 2px solid #CAB8FF;
                border-radius: 30px;
                font-weight: bold;
                font-size: 60px;
            }
            QPushButton:hover {
                background-color: #CAB8FF;
                color: white;
            }
        """)

        self.next_btn_rect = self.next_btn.geometry()
        self.change_sentence_btn_rect = self.change_sentence_btn.geometry()

        self.layout.addStretch()
        self.setLayout(self.layout)

        self.start_time = time.time()
        self.air_start_time = self.start_time

    def load_alzheimers_model(self):
        try:
            import joblib
            import numpy as np
            self.Model_FOLDER=r"C:\Users\Hooria\PycharmProjects\Project4\.venv"
            model_path = os.path.join(self.Model_FOLDER, "model14.pkl")
            scaler_path = os.path.join(self.Model_FOLDER, "scaler14.pkl")

            if os.path.exists(model_path) and os.path.exists(scaler_path):
                try:
                    self.alzheimers_model = joblib.load(model_path)
                    self.alzheimers_scaler = joblib.load(scaler_path)
                    self.alzheimers_model_loaded = True
                except Exception as e:
                    self.alzheimers_model = None
                    self.alzheimers_scaler = None
                    self.alzheimers_model_loaded = False
            else:
                self.alzheimers_model = None
                self.alzheimers_scaler = None
                self.alzheimers_model_loaded = False

        except ImportError as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False
        except Exception as e:
            self.alzheimers_model = None
            self.alzheimers_scaler = None
            self.alzheimers_model_loaded = False

    def calculate_drawing_metrics(self):
        if not self.pen_positions or not self.pen_timestamps:
            return None

        try:
            import numpy as np

            positions = np.array(self.pen_positions)
            timestamps = np.array(self.pen_timestamps)
            pressures = np.array(self.pressure_readings) if self.pressure_readings else np.ones(len(positions))

            total_time = self.end_time - self.start_time if self.start_time and self.end_time else 0
            air_time = self.air_time
            paper_time = self.paper_time

            velocities = []
            accelerations = []
            jerks = []

            if len(positions) > 1:
                for i in range(1, len(positions)):
                    dt = timestamps[i] - timestamps[i - 1]
                    if dt > 0:
                        dx = positions[i][0] - positions[i - 1][0]
                        dy = positions[i][1] - positions[i - 1][1]
                        velocity = np.sqrt(dx ** 2 + dy ** 2) / dt
                        velocities.append(velocity)

                        if len(velocities) > 1:
                            dv = velocities[-1] - velocities[-2]
                            acceleration = dv / dt
                            accelerations.append(acceleration)

                            if len(accelerations) > 1:
                                da = accelerations[-1] - accelerations[-2]
                                jerk = da / dt
                                jerks.append(jerk)

            mean_speed = np.mean(velocities) if velocities else 0
            mean_acc = np.mean(accelerations) if accelerations else 0
            mean_jerk = np.mean(jerks) if jerks else 0

            pressure_mean = np.mean(pressures) if len(pressures) > 0 else 0
            pressure_var = np.var(pressures) if len(pressures) > 0 else 0

            max_x = np.max(positions[:, 0]) - np.min(positions[:, 0]) if len(positions) > 0 else 0
            max_y = np.max(positions[:, 1]) - np.min(positions[:, 1]) if len(positions) > 0 else 0

            gmrtp = total_time / max(1, len(positions)) if len(positions) > 0 else 0

            if len(positions) > 1:
                center_x = np.mean(positions[:, 0])
                center_y = np.mean(positions[:, 1])
                distances = np.sqrt((positions[:, 0] - center_x) ** 2 + (positions[:, 1] - center_y) ** 2)
                disp_index = np.std(distances)
            else:
                disp_index = 0

            features = [
                total_time,
                air_time,
                paper_time,
                mean_speed,
                mean_acc,
                pressure_mean,
                pressure_var,
                self.pendown_count,
                max_x,
                max_y,
                gmrtp,
                mean_jerk,
                disp_index
            ]

            return features

        except Exception as e:
            return None

    def predict_alzheimers_risk(self, features):
        if not self.alzheimers_model_loaded or self.alzheimers_model is None or self.alzheimers_scaler is None:
            return {
                'risk_score': 0.0,
                'risk_level': 'Model Not Available',
                'interpretation': 'Alzheimer\'s detection model could not be loaded. Please check if model14.pkl and scaler14.pkl exist in the save folder.'
            }

        try:
            import numpy as np

            if len(features) != 13:
                while len(features) < 13:
                    features.append(0.0)
                features = features[:13]

            X = np.array(features).reshape(1, -1)

            X_scaled = self.alzheimers_scaler.transform(X)

            prediction_proba = self.alzheimers_model.predict_proba(X_scaled)

            risk_score = prediction_proba[0][1] if len(prediction_proba[0]) > 1 else prediction_proba[0][0]

            if risk_score <= 0.33:
                risk_level = "Low Risk"
                interpretation = "The drawing patterns suggest low risk for Alzheimer's-related cognitive decline."
            elif risk_score <= 0.66:
                risk_level = "Medium Risk"
                interpretation = "The drawing patterns suggest moderate risk. Consider follow-up assessment."
            else:
                risk_level = "High Risk"
                interpretation = "The drawing patterns suggest higher risk. Professional evaluation recommended."

            result = {
                'risk_score': float(risk_score),
                'risk_level': risk_level,
                'interpretation': interpretation
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': 'Prediction Error',
                'interpretation': f'Error during analysis: {str(e)}'
            }

    def run_alzheimers_analysis(self):
        features = self.calculate_drawing_metrics()

        if features is None:
            self.alzheimers_result = {
                'risk_score': 0.0,
                'risk_level': 'No Data',
                'interpretation': 'Insufficient drawing data for analysis'
            }
            return False

        self.alzheimers_result = self.predict_alzheimers_risk(features)
        return True

    def update_sentence_display(self):
        current_sentence = self.sentences[self.current_sentence_index]
        self.sentence_box.setText(current_sentence)

    def change_sentence(self):
        self.current_sentence_index = (self.current_sentence_index + 1) % len(self.sentences)
        self.update_sentence_display()

        if self.drawing:
            self.drawing.fill(Qt.darkCyan)
            self.draw_writing_line()
            self.update()

        if self.user_drawing:
            self.user_drawing.fill(Qt.transparent)

    def showEvent(self, event):
        super().showEvent(event)

        self.drawing = QPixmap(self.size())
        self.drawing.fill(QColor(140, 120, 204))

        self.user_drawing = QPixmap(self.size())
        self.user_drawing.fill(Qt.transparent)

        self.draw_writing_line()

    def draw_writing_line(self):
        if self.drawing is None:
            return

        painter = QPainter(self.drawing)
        pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
        painter.setPen(pen)

        line_y = self.height() - 300
        line_start_x = 50
        line_end_x = self.width() - 50

        painter.drawLine(line_start_x, line_y, line_end_x, line_y)

        font = QFont("Arial", 26)
        painter.setFont(font)
        painter.setPen(QPen(Qt.white, 2, Qt.SolidLine, Qt.RoundCap))
        painter.drawText(0, line_y - 20, "Write here:")

        painter.end()

    def is_over_buttons(self, pos):
        self.next_btn_rect = self.next_btn.geometry()
        self.change_sentence_btn_rect = self.change_sentence_btn.geometry()
        return (self.next_btn_rect.contains(pos) or
                self.change_sentence_btn_rect.contains(pos))

    def create_white_background_drawing(self):
        if self.user_drawing is None:
            return QPixmap(self.size())

        original_height = self.user_drawing.height()
        original_width = self.user_drawing.width()
        crop_top = 300
        crop_bottom = 200

        if crop_top + crop_bottom >= original_height:
            crop_top = 0
            crop_bottom = 0

        cropped_height = original_height - crop_top - crop_bottom

        white_drawing = QPixmap(original_width, cropped_height)
        white_drawing.fill(Qt.white)

        painter = QPainter(white_drawing)
        painter.drawPixmap(0, 0, self.user_drawing, 0, crop_top, original_width, cropped_height)
        painter.end()

        return white_drawing

    def handle_next(self):
        if self.end_time is None:
            self.end_time = time.time()

        if self.last_pen_up_time is not None:
            self.air_time += (self.end_time - self.last_pen_up_time)

        alzheimers_success = self.run_alzheimers_analysis()

        if self.has_drawing_content():
            self.save_image_and_log()
        else:
            self.save_basic_metrics()

        self.show_popup_with_home()

    def has_drawing_content(self):
        return len(self.pen_positions) > 0

    def save_basic_metrics(self):
        if self.metrics_logger is None:
            return

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            total_time = self.end_time - self.start_time if self.start_time else 0

        except Exception as e:
            pass

    def save_image_and_log(self):
        if self.metrics_logger is None:
            return

        if self.user_drawing is None:
            return

        try:
            if self.start_time is None:
                self.start_time = time.time()
            if self.end_time is None:
                self.end_time = time.time()

            if not self.pen_positions:
                self.save_basic_metrics()
                return

            white_drawing = self.create_white_background_drawing()

            current_sentence = self.sentences[self.current_sentence_index]
            level_identifier = f"Level 8"

            metrics = self.metrics_logger.save_complete_session(
                drawing_pixmap=white_drawing,
                player_name=self.player_name,
                level=level_identifier,
                pen_positions=self.pen_positions,
                pen_timestamps=self.pen_timestamps,
                pressure_readings=self.pressure_readings,
                start_time=self.start_time,
                end_time=self.end_time,
                air_time=self.air_time,
                paper_time=self.paper_time,
                pendown_count=self.pendown_count
            )
            return metrics

        except Exception as e:
            self.save_basic_metrics()
            return None

    def show_popup_with_home(self):
        popup = QDialog(self)
        popup.setWindowTitle("Analysis Complete")
        popup.setFixedSize(700, 800)

        layout = QVBoxLayout()

        label = QLabel("Excellent Writing!")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("""
            QLabel {
                color: Black;
                font-weight: bold;
                font-size: 50px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(label)

        if self.alzheimers_result:
            alz_title = QLabel("Alzheimer's Risk Assessment")
            alz_title.setWordWrap(True)
            alz_title.setAlignment(Qt.AlignCenter)
            alz_title.setStyleSheet("""
                QLabel {
                    color: #2E8B57;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_title)

            alz_risk_label = QLabel(f"Risk Level: {self.alzheimers_result.get('risk_level', 'Unknown')}")
            alz_risk_label.setWordWrap(True)
            alz_risk_label.setAlignment(Qt.AlignCenter)
            alz_risk_label.setStyleSheet("""
                QLabel {
                    color: Black;
                    font-weight: bold;
                    font-size: 25px;
                    margin-bottom: 5px;
                }
            """)
            layout.addWidget(alz_risk_label)


            alz_interpretation = QLabel(self.alzheimers_result.get('interpretation', 'No interpretation available'))
            alz_interpretation.setWordWrap(True)
            alz_interpretation.setAlignment(Qt.AlignCenter)
            alz_interpretation.setStyleSheet("""
                QLabel {
                    color: #444444;
                    font-size: 20px;
                    margin-bottom: 20px;
                    padding: 10px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: #f9f9f9;
                }
            """)
            layout.addWidget(alz_interpretation)

        home_btn = QPushButton("Home")
        home_btn.setFixedSize(200, 60)
        home_btn.clicked.connect(lambda: self.return_home_from_popup(popup))
        home_btn.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                background-color: #4169E1;
                border-radius: 15px;
                font-weight: bold;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: #6495ED;
            }
        """)
        layout.addWidget(home_btn, alignment=Qt.AlignCenter)
        popup.setLayout(layout)
        popup.exec_()

    def return_home_from_popup(self, popup):
        popup.accept()
        self.open_main_screen()

    def resizeEvent(self, event):
        super().resizeEvent(event)

        size = event.size()
        if size.width() > 0 and size.height() > 0:
            old_drawing = self.drawing
            old_user_drawing = self.user_drawing

            self.drawing = QPixmap(size)
            self.drawing.fill(Qt.darkCyan)

            self.user_drawing = QPixmap(size)
            self.user_drawing.fill(Qt.transparent)

            if old_drawing and not old_drawing.isNull():
                painter = QPainter(self.drawing)
                painter.drawPixmap(0, 0, old_drawing)
                painter.end()

            if old_user_drawing and not old_user_drawing.isNull():
                painter = QPainter(self.user_drawing)
                painter.drawPixmap(0, 0, old_user_drawing)
                painter.end()

            self.draw_writing_line()

    def paintEvent(self, event):
        if self.drawing is not None:
            base_painter = QPainter(self)
            base_painter.drawPixmap(0, 0, self.drawing)
            if self.user_drawing is not None:
                base_painter.drawPixmap(0, 0, self.user_drawing)

    def mousePressEvent(self, event):
        if self.is_over_buttons(event.pos()):
            return

        timestamp = time.time()

        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp
            self.air_start_time = timestamp

        if event.button() == Qt.LeftButton:
            if self.last_pen_up_time is not None:
                self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

            self.is_drawing = True
            self.last_point = event.pos()
            self.paper_start_time = timestamp
            self.last_time = timestamp

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(0.5)

    def mouseMoveEvent(self, event):
        if self.is_over_buttons(event.pos()):
            return

        if event.buttons() & Qt.LeftButton and self.last_point and self.is_drawing and self.user_drawing is not None:
            timestamp = time.time()

            painter = QPainter(self.user_drawing)
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(0.5)

            if self.last_time is not None:
                delta_time = timestamp - self.last_time
                self.paper_time += delta_time
            self.last_time = timestamp

    def mouseReleaseEvent(self, event):
        if self.is_drawing:
            self.is_drawing = False
            self.last_point = None
            self.last_pen_up_time = time.time()

    def tabletEvent(self, event: QTabletEvent):
        if self.is_over_buttons(event.pos()):
            return

        timestamp = time.time()

        if self.start_time is None:
            self.start_time = timestamp
            self.last_time = timestamp
            self.air_start_time = timestamp

        if event.type() == QTabletEvent.TabletPress:
            if self.last_pen_up_time is not None:
                self.air_time += (timestamp - self.last_pen_up_time)
                self.last_pen_up_time = None

            self.is_drawing = True
            self.pendown_count += 1
            self.last_point = event.pos()
            self.paper_start_time = timestamp
            self.last_time = timestamp

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(event.pressure())

        elif event.type() == QTabletEvent.TabletMove and self.last_point and self.is_drawing and self.user_drawing is not None:
            painter = QPainter(self.user_drawing)
            pen = QPen(Qt.black, 20, Qt.SolidLine, Qt.RoundCap)
            painter.setPen(pen)
            painter.drawLine(self.last_point, event.pos())
            painter.end()

            self.last_point = event.pos()
            self.update()

            self.pen_positions.append((event.pos().x(), event.pos().y()))
            self.pen_timestamps.append(timestamp)
            self.pressure_readings.append(event.pressure())

            if self.last_time is not None:
                delta_time = timestamp - self.last_time
                self.paper_time += delta_time
            self.last_time = timestamp

        elif event.type() == QTabletEvent.TabletRelease:
            if self.is_drawing:
                self.is_drawing = False
                self.last_point = None
                self.last_pen_up_time = time.time()

        event.accept()

    def open_main_screen(self):
        self.hide()
        self.main_dialog.show()


class Name_Screen(QWidget):
    def __init__(self):
        super().__init__()
        self.setMinimumSize(2000, 1100)
        self.resize(2000, 1100)

        # Background setup
        self.bg_label = QLabel(self)
        self.bg_label.setScaledContents(True)
        self.pixmap = QPixmap("Images/back2.jpg")
        if self.pixmap.isNull():
            print("Error: 'back2.jpg' not found or could not be loaded.")
        else:
            self.bg_label.setPixmap(self.pixmap)
            self.bg_label.resize(self.size())

        # Next button
        self.button = QPushButton("Next", self)
        self.button.resize(300, 150)
        self.button.move(1750, 950)
        self.button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
        """)
        self.button.clicked.connect(self.proceed_to_game_menu)

        # Back button
        self.button1 = QPushButton("Back", self)
        self.button1.resize(200, 100)
        self.button1.move(0, 0)
        self.button1.setStyleSheet("""
            QPushButton {
            background-color: black;
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: bold;
                font-size: 50px;
            }
        """)
        self.button1.clicked.connect(self.back_button)

        layout = QVBoxLayout()

        # Title
        title = QLabel("Please enter your name to begin your adventure!", self)
        title.setFont(QFont("Arial", 32, QFont.Bold))
        title.setStyleSheet("background-color: black; color: #ecf0f1;font-weight: bold;")
        title.move(10, 340)
        layout.addWidget(title)

        # Name input
        input_layout = QHBoxLayout()
        input_widget = QWidget()
        input_widget.setLayout(input_layout)

        name_label = QLabel("Name:", self)
        name_label.setFont(QFont("Arial", 25))
        name_label.setStyleSheet("background-color: black;font-weight: bold; color: #ecf0f1; margin: 30px;")
        name_label.resize(290, 150)
        name_label.move(310, 425)
        layout.addWidget(name_label)

        self.name_input = QLineEdit(self)
        self.name_input.setFont(QFont("Arial", 25))
        self.name_input.setFixedHeight(100)
        self.name_input.setFixedWidth(1000)
        self.name_input.setStyleSheet("""
                QLineEdit {
                    padding: 10px;
                    border: 2px solid #3498db;
                    border-radius: 20px;
                    background-color: black;
                    font-weight: bold;
                    color: white;
                }
                QLineEdit:focus {
                    border: 2px solid #2ecc71;
                    background-color: black;
                }
            """)
        self.name_input.setPlaceholderText("Enter your name here...")
        self.name_input.textChanged.connect(self.check_input)
        self.name_input.returnPressed.connect(self.proceed_to_game_menu)
        self.name_input.move(580, 450)
        layout.addWidget(self.name_input)

    def resizeEvent(self, event):
        if not self.pixmap.isNull():
            self.bg_label.setPixmap(self.pixmap.scaled(
                self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation
            ))
            self.bg_label.resize(self.size())

    def check_input(self):
        self.button.setEnabled(bool(self.name_input.text().strip()))

    def proceed_to_game_menu(self):
        if self.name_input.text().strip():
            self.hide()
            self.dialog = QDialog()
            self.ui = Ui_Dialog()
            self.ui.setupUi(self.dialog, self.name_input.text().strip())
            self.dialog.exec_()

    def back_button(self):
        self.hide()
        self.main_menu = MainMenuScreen()
        self.main_menu.show()


class MainMenuScreen(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Welcome to AI: Aging Intelligently")
        self.setMinimumSize(2000, 1100)
        self.resize(2000, 1100)

        self.bg1_label = QLabel(self)
        self.bg1_label.setScaledContents(True)
        self.original_pixmap = QPixmap("Images/back.png")
        if self.original_pixmap.isNull():
            print("Error: 'back.png' not found or could not be loaded.")
        else:
            self.bg1_label.setPixmap(self.original_pixmap)
            self.bg1_label.resize(self.size())

        # Title Label
        self.Tlabel = QLabel("Welcome to\nAI: Aging Intelligently", self)
        self.Tlabel.setFont(QFont("Fantasy", 60))
        self.Tlabel.setStyleSheet("background-color: black;color: white;font-weight: bold;")
        self.Tlabel.setAlignment(Qt.AlignCenter)
        self.Tlabel.setMinimumSize(2000,400)
        self.Tlabel.move(0,150)
        # Next button
        self.button = QPushButton("Next", self)
        self.button.resize(300, 150)
        self.button.move(1750, 950)
        self.button.setStyleSheet("""
            QPushButton {
                color: white;
                border: none;
                border-radius: 30px;
                font-weight: bold;
                font-size: 50px;
            }
        """)
        self.button.clicked.connect(self.open_name_screen)

    def resizeEvent(self, event):
        if not self.original_pixmap.isNull():
            self.bg1_label.setPixmap(self.original_pixmap.scaled(
                self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation
            ))
            self.bg1_label.resize(self.size())
        super().resizeEvent(event)

    def open_name_screen(self):
        self.hide()
        self.name_screen = Name_Screen()
        self.name_screen.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainMenuScreen()
    window.show()
    sys.exit(app.exec_())
