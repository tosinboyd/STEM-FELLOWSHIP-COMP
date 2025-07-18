import os
import numpy as np
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras.applications import MobileNetV2
from tensorflow.keras.optimizers import Adam
from PIL import Image, ImageEnhance, ImageFilter
import warnings
import openpyxl
from openpyxl import load_workbook

warnings.filterwarnings('ignore')

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
tf.config.threading.set_inter_op_parallelism_threads(0)
tf.config.threading.set_intra_op_parallelism_threads(0)

try:
    tf.config.optimizer.set_jit(True)
except:
    pass

gpus = tf.config.list_physical_devices('GPU')
if gpus:
    try:
        for gpu in gpus:
            tf.config.experimental.set_memory_growth(gpu, True)
        tf.keras.mixed_precision.set_global_policy('mixed_float16')
    except RuntimeError:
        pass
else:
    tf.keras.mixed_precision.set_global_policy('float32')

np.random.seed(42)
tf.random.set_seed(42)


class ImagePreprocessor:

    @staticmethod
    def remove_background_and_isolate_drawing(image_path, output_path=None):
        try:
            image = Image.open(image_path)
            if image is None:
                raise ValueError(f"Could not load image from {image_path}")

            if image.mode != 'RGB':
                image = image.convert('RGB')

            gray_image = image.convert('L')
            blurred = gray_image.filter(ImageFilter.GaussianBlur(radius=1.5))
            gray_array = np.array(blurred)

            hist, _ = np.histogram(gray_array.flatten(), bins=256, range=(0, 256))

            total_pixels = gray_array.size
            current_max = 0
            threshold = 0
            sum_total = np.sum(np.arange(256) * hist)
            sum_background = 0
            weight_background = 0

            for i in range(256):
                weight_background += hist[i]
                if weight_background == 0:
                    continue

                weight_foreground = total_pixels - weight_background
                if weight_foreground == 0:
                    break

                sum_background += i * hist[i]
                mean_background = sum_background / weight_background
                mean_foreground = (sum_total - sum_background) / weight_foreground

                between_class_variance = weight_background * weight_foreground * (
                        mean_background - mean_foreground) ** 2

                if between_class_variance > current_max:
                    current_max = between_class_variance
                    threshold = i

            binary_array = (gray_array > threshold).astype(np.uint8) * 255
            binary_image = Image.fromarray(binary_array, mode='L')
            cleaned = binary_image.filter(ImageFilter.MedianFilter(size=3))
            edge_enhanced = cleaned.filter(ImageFilter.EDGE_ENHANCE_MORE)

            cleaned_array = np.array(cleaned)
            edge_array = np.array(edge_enhanced)

            combined_array = np.logical_or(cleaned_array > 128, edge_array > 128).astype(np.uint8) * 255
            inverted_array = 255 - combined_array
            final_image_pil = Image.fromarray(inverted_array, mode='L')
            final_image_rgb = final_image_pil.convert('RGB')
            final_image_array = np.array(final_image_rgb)

            if output_path:
                final_image_rgb.save(output_path)

            return final_image_array

        except Exception as e:
            return None

    @staticmethod
    def enhance_line_drawing(image_array):
        try:
            if image_array.max() <= 1.0:
                image_array = (image_array * 255).astype(np.uint8)

            pil_image = Image.fromarray(image_array)
            enhancer = ImageEnhance.Contrast(pil_image)
            enhanced = enhancer.enhance(2.0)
            sharpness_enhancer = ImageEnhance.Sharpness(enhanced)
            sharpened = sharpness_enhancer.enhance(1.5)
            edge_enhanced = sharpened.filter(ImageFilter.EDGE_ENHANCE)
            unsharp_enhanced = edge_enhanced.filter(ImageFilter.UnsharpMask(radius=1, percent=150, threshold=3))
            enhanced_array = np.array(unsharp_enhanced)

            return enhanced_array

        except Exception as e:
            return image_array


class ParkinsonsDetector:

    def __init__(self, model_path='best_parkinsons_model.keras', image_size=(128, 128)):
        self.model_path = model_path
        self.image_size = image_size
        self.model = None
        self.preprocessor = ImagePreprocessor()
        self.class_names = ['healthy', 'parkinson']

        self.load_model()

    def load_model(self):
        if os.path.exists(self.model_path):
            try:
                self.model = keras.models.load_model(self.model_path)
                return True
            except Exception as e:
                return False
        return False

    def preprocess_image(self, image_path):
        try:
            processed_image = self.preprocessor.remove_background_and_isolate_drawing(image_path)

            if processed_image is None:
                return None

            enhanced_image = self.preprocessor.enhance_line_drawing(processed_image)

            image_tensor = tf.constant(enhanced_image, dtype=tf.float32)
            image_tensor = tf.image.resize(image_tensor, self.image_size, method='bilinear')
            image_tensor = image_tensor / 255.0
            image_tensor = tf.expand_dims(image_tensor, 0)

            return image_tensor

        except Exception as e:
            return None

    def predict(self, image_path):
        if self.model is None:
            return None

        processed_image = self.preprocess_image(image_path)
        if processed_image is None:
            return None

        try:
            prediction = self.model.predict(processed_image, verbose=0)[0][0]
            risk_score = float(prediction)

            if risk_score < 0.35:
                risk_level = "Low Risk"
                interpretation = "Spiral drawing shows characteristics typical of healthy motor control"
            elif risk_score < 0.65:
                risk_level = "Moderate Risk"
                interpretation = "Spiral drawing shows some irregularities that may warrant further evaluation"
            else:
                risk_level = "High Risk"
                interpretation = "Spiral drawing shows significant irregularities consistent with motor control issues"

            distance_from_center = abs(risk_score - 0.5)
            confidence = min(distance_from_center * 2, 0.95)

            result = {
                'risk_score': risk_score,
                'risk_level': risk_level,
                'confidence': confidence,
                'interpretation': interpretation,
                'prediction_successful': True
            }

            return result

        except Exception as e:
            return {
                'risk_score': 0.0,
                'risk_level': "Error",
                'confidence': 0.0,
                'interpretation': "Unable to process image",
                'prediction_successful': False
            }

    def update_excel_with_results(self, excel_path, player_name, risk_level, evaluation_results):
        try:
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                return False

            player_row = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == player_name:
                    player_row = row
                    break

            if player_row is None:
                return False

            ws.cell(row=player_row, column=18, value=risk_level)

            if evaluation_results:
                eval_text = f"Acc: {evaluation_results.get('accuracy', 0):.3f}, Prec: {evaluation_results.get('precision', 0):.3f}, Rec: {evaluation_results.get('recall', 0):.3f}, F1: {evaluation_results.get('f1_score', 0):.3f}, AUC: {evaluation_results.get('auc', 0):.3f}"
            else:
                eval_text = "Model evaluation not available"

            ws.cell(row=player_row, column=19, value=eval_text)

            wb.save(excel_path)
            return True

        except Exception as e:
            return False

    def cleanup_processed_image(self, image_path):
        try:
            if os.path.exists(image_path):
                os.remove(image_path)
        except Exception as e:
            print(f"Error deleting processed image: {e}")

